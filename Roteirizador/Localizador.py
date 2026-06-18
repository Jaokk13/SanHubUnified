import sys
import os
import io
import json
import math
import urllib.parse
import folium
import unicodedata
import difflib
import requests
import time
import re
import pandas as pd
import openpyxl

# Tenta importar qtawesome, usa fallback se não existir (evita erro ao abrir)
try:
    import qtawesome as qta
except ImportError:
    class qta:
        @staticmethod
        def icon(*args, **kwargs): return QIcon()

import shutil
from geopy.geocoders import Nominatim, ArcGIS
from geopy.extra.rate_limiter import RateLimiter

# Imports da Interface Gráfica (PyQt5)
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QTextEdit, QPushButton, QLabel, QProgressBar, QMessageBox, QLineEdit, QFileDialog, QComboBox, QDialog, QFormLayout, QListWidget, QListWidgetItem, QAbstractItemView, QFrame, QToolTip, QSpinBox, QInputDialog)
from PyQt5.QtWebEngineWidgets import QWebEngineView, QWebEnginePage
from PyQt5.QtCore import QThread, pyqtSignal, Qt
from PyQt5.QtGui import QIcon

# --- CONFIGURAÇÃO GLOBAL ---
APP_CONFIG = {
    "cidade": "Cuiabá",
    "estado": "MT",
    "base_address": "-15.635946308016816, -56.02728144730883",
    "bbox": [-15.85, -15.35, -56.35, -55.75], # [min_lat, max_lat, min_lon, max_lon]
    "center": [-15.5989, -56.0949]
}

# Define o diretório base (Compatível com .EXE e Script Python)
if getattr(sys, 'frozen', False):
    SCRIPT_DIR = os.path.dirname(sys.executable)
else:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# --- CONFIGURAÇÃO DE DIRETÓRIOS (APPDATA) ---
# Para funcionar em "Arquivos de Programas", os dados devem ficar em AppData
APP_NAME = "RoteirizadorBrPaving"
DATA_DIR = os.path.join(os.getenv('APPDATA'), APP_NAME)

if not os.path.exists(DATA_DIR):
    os.makedirs(DATA_DIR)

# Arquivos agora ficam na pasta de dados do usuário
DB_FILE = os.path.join(DATA_DIR, "banco_bairros_oficial.json")
SETTINGS_FILE = os.path.join(DATA_DIR, "window_cache.json")
CACHE_FILE = os.path.join(DATA_DIR, "enderecos_cache.json")
CONFIG_FILE = os.path.join(DATA_DIR, "app_config.json")

def load_config():
    global APP_CONFIG
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                saved = json.load(f)
                APP_CONFIG.update(saved)
            
            # Validação de segurança para o BBox (evita crash se estiver corrompido/null)
            bbox = APP_CONFIG.get('bbox')
            if not isinstance(bbox, list) or len(bbox) != 4:
                APP_CONFIG['bbox'] = [-90, 90, -180, 180] # Reset para o mundo todo
        except: pass

def save_config():
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(APP_CONFIG, f, ensure_ascii=False, indent=4)
    except: pass

# --- FUNÇÕES AUXILIARES (Lógica Pura) ---

def identificar_zona(lat, lon):
    import math
    # Centro dinâmico baseado na configuração
    center = APP_CONFIG.get('center', [-15.5989, -56.0949])
    diff_lat = lat - center[0]
    diff_lon = lon - center[1]
    
    # math.atan2(y, x) onde y=diff_lon e x=diff_lat. Converte para graus.
    angle = math.degrees(math.atan2(diff_lon, diff_lat))
    if angle < 0: 
        angle += 360
        
    # Divide os 360 graus em 8 fatias de 45 graus, centradas nos eixos cardeais
    # Norte: 337.5 a 22.5
    if angle >= 337.5 or angle < 22.5: return "Zona Norte"
    if angle >= 22.5 and angle < 67.5: return "Zona Nordeste"
    if angle >= 67.5 and angle < 112.5: return "Zona Leste"
    if angle >= 112.5 and angle < 157.5: return "Zona Sudeste"
    if angle >= 157.5 and angle < 202.5: return "Zona Sul"
    if angle >= 202.5 and angle < 247.5: return "Zona Sudoeste"
    if angle >= 247.5 and angle < 292.5: return "Zona Oeste"
    if angle >= 292.5 and angle < 337.5: return "Zona Noroeste"

def calcular_distancia(coord1, coord2):
    # Fórmula de Haversine para distância em linha reta (aproximada)
    R = 6371  # Raio da Terra em km
    lat1, lon1 = math.radians(coord1[0]), math.radians(coord1[1])
    lat2, lon2 = math.radians(coord2[0]), math.radians(coord2[1])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = math.sin(dlat / 2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c

def normalizar_endereco(texto):
    # Expande abreviações comuns para melhorar a precisão da busca
    # Substitui pontos por espaços para evitar problemas com "JD.Brasil" (sem espaço) ou "Av.CPA"
    texto = texto.replace('.', ' ')
    
    abreviacoes = {
        "adm": "Administrativo",
        "res": "Residencial",
        "jd": "Jardim",
        "jdm": "Jardim",
        "pq": "Parque",
        "vl": "Vila",
        "av": "Avenida",
        "dr": "Doutor",
        "sta": "Santa",
        "sto": "Santo",
        "n": "Nossa",
        "sra": "Senhora",
        "prof": "Professor"
    }
    palavras = texto.split()
    return " ".join([abreviacoes.get(p.lower(), p) for p in palavras])

def remover_acentos(texto):
    # Remove acentos para comparação de strings (ex: Cuiabá -> Cuiaba)
    nfkd = unicodedata.normalize('NFKD', texto)
    return "".join([c for c in nfkd if not unicodedata.combining(c)])

def validar_coordenadas_raw(lat, lon):
    """Valida se as coordenadas estão dentro dos limites (BBox e Raio) da cidade."""
    if lat is None or lon is None: return False
    
    # 1. Validação por BBox (Caixa)
    bbox = APP_CONFIG.get('bbox')
    if bbox and len(bbox) == 4:
        lat_min, lat_max, lon_min, lon_max = bbox
        if not (lat_min <= lat <= lat_max and lon_min <= lon <= lon_max):
            return False

    # 2. Validação por Distância do Centro (Raio de Segurança)
    # O usuário pediu para limitar a KM do ponto inicial. Usamos o centro da cidade como referência segura.
    center = APP_CONFIG.get('center')
    if center and len(center) == 2:
        try:
            dist = calcular_distancia((lat, lon), (center[0], center[1]))
            if dist > 100: # Limite rígido de 100km
                return False
        except: pass
        
    return True

def validar_localizacao(loc, ignorar_pois=False, permitir_condominio=False):
    """Valida se a localização encontrada é aceitável para a cidade configurada."""
    if not loc: return False
    
    # 1. Validação por Coordenadas (Filtro Geográfico Rigoroso e Raio)
    if not validar_coordenadas_raw(loc.latitude, loc.longitude):
        return False

    # 2. Verifica se é a cidade configurada (Texto)
    addr_clean = remover_acentos(loc.address.lower())
    cidade_alvo = remover_acentos(APP_CONFIG.get('cidade', '')).lower()
    
    if cidade_alvo and cidade_alvo not in addr_clean:
        return False
    
    # --- NOVO: Filtro de Condomínios ---
    # Se o usuário NÃO digitou "cond" na busca, rejeitamos resultados que sejam condomínios.
    if not permitir_condominio and "condominio" in addr_clean:
        return False

    # --- NOVO: Filtro de Resultado Genérico (ArcGIS) ---
    # ArcGIS costuma retornar apenas "Cuiabá, Mato Grosso" quando não encontra o endereço específico.
    # Isso joga o ponto para o centro da cidade. Vamos filtrar isso para evitar falsos positivos.
    # Gera strings genéricas baseadas na cidade atual
    genericos = [f"{cidade_alvo}, {remover_acentos(APP_CONFIG.get('estado', ''))}", f"{cidade_alvo}", f"{cidade_alvo}, brasil"]
    if any(g in addr_clean.strip() for g in genericos) and len(addr_clean) < len(cidade_alvo) + 20:
        return False

    # 3. Verifica Rank (Evita pino no centro da cidade - Rank <= 16)
    rank = loc.raw.get('place_rank')
    if rank and int(rank) <= 16:
        return False
    
    loc_class = loc.raw.get('class', '')
    
    # BLOQUEIO ABSOLUTO DE RIOS E NATUREZA (Correção para o Rio Três Barras)
    if loc_class in ['waterway', 'natural']:
        return False
    
    # 4. Filtro de POIs (Evita que bairros sejam confundidos com lojas/bancos)
    if ignorar_pois:
        # Classes inválidas para bairros (comércio, amenidades, etc.)
        invalid_classes = ['amenity', 'shop', 'tourism', 'leisure', 'office', 'craft', 'man_made']
        if loc_class in invalid_classes:
            return False
            
    # 5. Filtro Anti-Rio (Heurística de nome para Cuiabá)
    # Se o nome do local começar com "Rio ", provavelmente é um rio e não o bairro.
    if loc.address and loc.address.strip().startswith("Rio "):
        return False
            
    return True

def eh_bairro_de_verdade(location, nome_buscado):
    """
    Filtra resultados que são ruas ou locais errados.
    Retorna True se for um bairro/cidade/distrito.
    Retorna False se for rua ou comércio.
    """
    if not location: return False
    
    # --- 1. Análise Específica do Nominatim (OSM) ---
    if hasattr(location, 'raw') and isinstance(location.raw, dict) and 'class' in location.raw:
        raw = location.raw
        classe = raw.get('class', '')
        tipo = raw.get('type', '')
        address = raw.get('address', {})
        
        # Detecta se o usuário buscou uma rua
        buscou_rua = any(ind in nome_buscado.lower() for ind in ['rua', 'avenida', 'travessa', 'alameda', 'beco', 'servidao', 'estrada', 'rodovia', 'av.'])

        # Rejeita vias de trânsito (exceto se o usuário buscou explicitamente uma rua)
        if classe == 'highway' and not buscou_rua:
            return False
            
        # Se tem 'road' no endereço estruturado, mas o tipo não é de localidade
        if 'road' in address:
            tipos_aceitos = ['neighbourhood', 'suburb', 'quarter', 'hamlet', 'borough', 'village', 'city', 'town', 'administrative', 'residential', 'place']
            if tipo not in tipos_aceitos:
                return False

    # --- 2. Análise Genérica de Texto (Serve para ArcGIS e Nominatim) ---
    addr_lower = location.address.lower() if location.address else ""
    nome_buscado_lower = nome_buscado.lower()
    
    # Palavras-chave
    indicadores_bairro = ['jardim', 'parque', 'residencial', 'bairro', 'setor', 'loteamento', 'condominio', 'vila']
    indicadores_rua = ['rua', 'avenida', 'travessa', 'alameda', 'beco', 'servidao', 'estrada', 'rodovia']
    
    # Se o usuário buscou explicitamente um bairro (ex: "Jardim X")
    indicador_usado = next((ind for ind in indicadores_bairro if ind in nome_buscado_lower), None)
    
    if indicador_usado:
        # Se o resultado começa com Rua/Av (indicando ser a via principal do resultado)
        # E o indicador do bairro (ex: "jardim") NÃO aparece no endereço retornado
        # Então provavelmente pegou a rua com mesmo nome.
        comeca_com_rua = any(addr_lower.strip().startswith(ind) for ind in indicadores_rua)
        
        if comeca_com_rua and indicador_usado not in addr_lower:
            return False

    return True

def organizar_rota_tsp(dados_locais):
    """
    Otimiza a rota usando Nearest Neighbor + 2-Opt.
    dados_locais: Lista de dicionários {'lat': float, 'lon': float, 'nome': str, 'tipo': str}
    """
    if not dados_locais: return []
    
    # --- 1. Nearest Neighbor (Heurística Inicial) ---
    rota_ordenada = [dados_locais[0]]
    
    visitados = {0} # Conjunto de índices visitados
    atual_idx = 0
    
    while len(rota_ordenada) < len(dados_locais):
        melhor_dist = float('inf')
        proximo_idx = -1
        
        for i in range(len(dados_locais)):
            if i not in visitados:
                p1 = (dados_locais[atual_idx]['lat'], dados_locais[atual_idx]['lon'])
                p2 = (dados_locais[i]['lat'], dados_locais[i]['lon'])
                dist = calcular_distancia(p1, p2)
                if dist < melhor_dist:
                    melhor_dist = dist
                    proximo_idx = i
        
        if proximo_idx != -1:
            visitados.add(proximo_idx)
            rota_ordenada.append(dados_locais[proximo_idx])
            atual_idx = proximo_idx
    
    # --- 2. 2-Opt (Otimização para remover cruzamentos) ---
    # Refina a rota trocando arestas para diminuir a distância total e desembaraçar o caminho
    
    def calcular_total(rota):
        d = 0
        for k in range(len(rota) - 1):
            p1 = (rota[k]['lat'], rota[k]['lon'])
            p2 = (rota[k+1]['lat'], rota[k+1]['lon'])
            d += calcular_distancia(p1, p2)
        # Considera o retorno à base (ciclo fechado) para otimização
        p_fim = (rota[-1]['lat'], rota[-1]['lon'])
        p_inicio = (rota[0]['lat'], rota[0]['lon'])
        d += calcular_distancia(p_fim, p_inicio)
        return d

    melhor_distancia = calcular_total(rota_ordenada)
    melhorou = True
    
    while melhorou:
        melhorou = False
        for i in range(1, len(rota_ordenada) - 1):
            for j in range(i + 1, len(rota_ordenada)):
                # Cria nova rota invertendo o segmento [i, j]
                nova_rota = rota_ordenada[:]
                nova_rota[i:j+1] = reversed(nova_rota[i:j+1])
                
                nova_distancia = calcular_total(nova_rota)
                
                if nova_distancia < melhor_distancia:
                    rota_ordenada = nova_rota
                    melhor_distancia = nova_distancia
                    melhorou = True
                    
    return rota_ordenada

def dividir_em_equipes(dados_locais, num_equipes, base=None):
    """
    Divide serviços em N equipes usando Algoritmo Sweep (Varredura Angular).
    Garante regiões que não se cruzam (fatias de pizza) e divisão exata de serviços.
    """
    import math
    
    if not dados_locais or num_equipes <= 0:
        return {}
    
    # Se há menos serviços que equipes, cada serviço vira uma equipe
    if num_equipes >= len(dados_locais):
        resultado = {}
        for i, d in enumerate(dados_locais):
            membros = [d]
            if base:
                membros = [base.copy()] + membros
                p_ret = base.copy()
                p_ret['nome'] += " (Retorno)"
                membros.append(p_ret)
            resultado[f"Equipe {i+1}"] = membros
        return resultado
    
    # --- 1. Algoritmo Sweep (Varredura Angular por Região) ---
    # Define o centro de referência ESTRITAMENTE pelo centro da cidade configurado
    # Isso garante que as "fatias" cruzem a cidade do centro para as bordas de forma uniforme.
    center = APP_CONFIG.get('center', [-15.5989, -56.0949])
    center_lat = center[0]
    center_lon = center[1]
        
    # Calcula o ângulo de cada ponto em relação ao centro da cidade
    dados_com_angulo = []
    for d in dados_locais:
        # math.atan2(y, x) onde y é a longitude e x a latitude
        angulo = math.atan2(d['lon'] - center_lon, d['lat'] - center_lat)
        dados_com_angulo.append((angulo, d))
        
    # Ordena pelos ângulos (cria as "fatias" da região)
    dados_com_angulo.sort(key=lambda x: x[0])
    dados_ordenados = [item[1] for item in dados_com_angulo]
    
    # --- 2. Divisão e Balanceamento Exatos ---
    resultado = {}
    total = len(dados_ordenados)
    
    tamanho_base = total // num_equipes
    sobra = total % num_equipes
    
    idx_atual = 0
    for j in range(num_equipes):
        # Distribui a sobra 1 a 1 para as primeiras equipes
        qtd = tamanho_base + (1 if j < sobra else 0)
        membros = dados_ordenados[idx_atual : idx_atual + qtd]
        idx_atual += qtd
        
        if not membros:
            continue
            
        # Adiciona base no início para o TSP
        if base:
            membros_com_base = [base.copy()] + membros
        else:
            membros_com_base = membros
        
        # Otimiza rota interna da equipe (Caixeiro Viajante)
        rota = organizar_rota_tsp(membros_com_base)
        
        # Fecha ciclo (retorno à base)
        if rota and base:
            p_retorno = base.copy()
            p_retorno['nome'] += " (Retorno)"
            rota.append(p_retorno)
            
        resultado[f"Equipe {j+1}"] = rota
        
    return resultado

# --- CLASSES AUXILIARES DE UI ---

class CustomWebEnginePage(QWebEnginePage):
    """Página Web personalizada para interceptar cliques em links especiais."""
    def acceptNavigationRequest(self, url, _type, isMainFrame):
        if url.scheme() == 'addbairro':
            # Decodifica o nome do bairro da URL (ex: addbairro:Jardim%20X -> Jardim X)
            bairro_nome = urllib.parse.unquote(url.path())
            
            # Chama o callback na janela principal se existir
            if hasattr(self, 'callback_add_bairro'):
                self.callback_add_bairro(bairro_nome)
            return False
        return super().acceptNavigationRequest(url, _type, isMainFrame)

class AddBairroDialog(QDialog):
    """Janela para adicionar bairro manualmente."""
    def __init__(self, nome_bairro, lat=None, lon=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Editar / Adicionar Local")
        self.setFixedWidth(400)
        # Estilo consistente com o app
        self.setStyleSheet("""
            QDialog { background-color: white; }
            QLabel { font-weight: bold; color: #333; font-family: 'Segoe UI'; }
            QLineEdit { padding: 6px; border: 1px solid #ccc; border-radius: 4px; background: #f9f9f9; }
            QPushButton { padding: 8px 15px; border-radius: 4px; font-weight: bold; font-family: 'Segoe UI'; }
            QPushButton[text="Salvar"] { background-color: #004aad; color: white; border: none; }
            QPushButton[text="Salvar"]:hover { background-color: #003380; }
            QPushButton[text="Cancelar"] { background-color: #d9534f; color: white; border: none; }
            QPushButton[text="Cancelar"]:hover { background-color: #c9302c; }
        """)
        
        layout = QVBoxLayout(self)
        form_layout = QFormLayout()
        form_layout.setSpacing(10)
        
        self.txt_nome = QLineEdit(nome_bairro)
        form_layout.addRow("Nome do Bairro:", self.txt_nome)
        
        self.txt_lat = QLineEdit(str(lat) if lat is not None else "")
        self.txt_lat.setPlaceholderText("Ex: -15.123456")
        form_layout.addRow("Latitude:", self.txt_lat)
        
        self.txt_lon = QLineEdit(str(lon) if lon is not None else "")
        self.txt_lon.setPlaceholderText("Ex: -56.123456")
        form_layout.addRow("Longitude:", self.txt_lon)
        
        layout.addLayout(form_layout)
        
        btn_layout = QHBoxLayout()
        btn_cancel = QPushButton("Cancelar")
        btn_cancel.clicked.connect(self.reject)
        
        btn_add = QPushButton("Salvar")
        btn_add.clicked.connect(self.validar_e_aceitar)
        
        btn_layout.addStretch()
        btn_layout.addWidget(btn_cancel)
        btn_layout.addWidget(btn_add)
        
        layout.addLayout(btn_layout)
        
    def validar_e_aceitar(self):
        try:
            # Valida conversão para float
            lat = float(self.txt_lat.text().replace(',', '.'))
            lon = float(self.txt_lon.text().replace(',', '.'))
            
            if not self.txt_nome.text().strip():
                QMessageBox.warning(self, "Erro", "O nome do bairro é obrigatório.")
                return
                
            self.accept()
        except ValueError:
            QMessageBox.warning(self, "Erro", "Coordenadas inválidas. Certifique-se de usar apenas números e ponto/vírgula.")

    def get_data(self):
        return {
            "nome": self.txt_nome.text().strip(),
            "lat": float(self.txt_lat.text().replace(',', '.')),
            "lon": float(self.txt_lon.text().replace(',', '.'))
        }

class ConfigDialog(QDialog):
    """Janela para configurar a cidade e base."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Configuração Inicial")
        self.setFixedWidth(400)
        self.setStyleSheet("QDialog { background-color: white; } QLabel { font-weight: bold; }")
        
        layout = QVBoxLayout(self)
        form = QFormLayout()
        
        self.txt_cidade = QLineEdit(APP_CONFIG.get('cidade', ''))
        self.txt_estado = QLineEdit(APP_CONFIG.get('estado', ''))
        self.txt_base = QLineEdit(APP_CONFIG.get('base_address', ''))
        
        form.addRow("Cidade:", self.txt_cidade)
        form.addRow("Estado (UF):", self.txt_estado)
        form.addRow("Endereço da Base:", self.txt_base)
        
        layout.addLayout(form)
        
        btn_save = QPushButton("Salvar e Configurar")
        btn_save.setStyleSheet("background-color: #004aad; color: white; padding: 10px; font-weight: bold;")
        btn_save.clicked.connect(self.salvar)
        layout.addWidget(btn_save)

    def salvar(self):
        cidade = self.txt_cidade.text().strip()
        estado = self.txt_estado.text().strip()
        base = self.txt_base.text().strip()
        
        if not cidade or not estado:
            QMessageBox.warning(self, "Erro", "Cidade e Estado são obrigatórios.")
            return
            
        # Busca dados da cidade
        try:
            geolocator = Nominatim(user_agent="roteirizador_config")
            loc = geolocator.geocode(f"{cidade}, {estado}", exactly_one=True)
            if loc:
                # Bounding Box [min_lat, max_lat, min_lon, max_lon]
                raw_bbox = loc.raw.get('boundingbox', [])
                bbox = [float(x) for x in raw_bbox] if raw_bbox else []
                
                APP_CONFIG['cidade'] = cidade
                APP_CONFIG['estado'] = estado
                APP_CONFIG['base_address'] = base
                APP_CONFIG['center'] = [loc.latitude, loc.longitude]
                
                if bbox and len(bbox) == 4: 
                    APP_CONFIG['bbox'] = bbox
                else:
                    APP_CONFIG['bbox'] = [-90, 90, -180, 180] # Reseta se não achar bbox da cidade
                
                save_config()
                self.accept()
            else:
                QMessageBox.warning(self, "Erro", "Cidade não encontrada no mapa. Verifique a grafia.")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao buscar cidade: {e}")

class MapGenWorker(QThread):
    """Worker dedicado apenas para gerar o HTML do mapa (evita travar a UI ao inverter)."""
    finished = pyqtSignal(str)
    
    def __init__(self, app_ref, rota):
        super().__init__()
        self.app = app_ref
        self.rota = rota
        
    def run(self):
        # Executa a geração do mapa (que contem chamadas de rede OSRM) em background
        html = self.app.gerar_mapa_html(self.rota)
        self.finished.emit(html)

class DatabaseUpdater(QThread):
    """Worker que atualiza o banco de dados via OSM ao iniciar."""
    finished = pyqtSignal(str)

    def run(self):
        try:
            # URL e Query do Overpass API (Dinâmico)
            overpass_url = "https://overpass-api.de/api/interpreter"
            cidade = APP_CONFIG.get('cidade', 'Cuiabá')
            overpass_query = f"""
            [out:json][timeout:180];
            area["name"="{cidade}"]["admin_level"="8"]->.searchArea;
            (
              node["place"~"suburb|neighbourhood|quarter"](area.searchArea);
              way["place"~"suburb|neighbourhood|quarter"](area.searchArea);
              rel["place"~"suburb|neighbourhood|quarter"](area.searchArea);
            );
            out center;
            """
            
            response = requests.get(overpass_url, params={'data': overpass_query}, timeout=60)
            if response.status_code == 200:
                data = response.json()
                
                # Carrega banco existente para não perder edições manuais
                db_existente = {}
                if os.path.exists(DB_FILE):
                    with open(DB_FILE, 'r', encoding='utf-8') as f:
                        db_existente = json.load(f)

                count = 0
                for element in data.get('elements', []):
                    if 'tags' in element and 'name' in element['tags']:
                        nome = element['tags']['name']
                        # Só atualiza se NÃO for manual (preserva edições do usuário)
                        if nome in db_existente and db_existente[nome].get('tipo') == 'manual':
                            continue
                        
                        lat = element.get('lat') or element.get('center', {}).get('lat')
                        lon = element.get('lon') or element.get('center', {}).get('lon')
                        
                        if lat and lon:
                            db_existente[nome] = {"lat": lat, "lon": lon, "tipo": "osm_auto"}
                            count += 1
                
                with open(DB_FILE, 'w', encoding='utf-8') as f:
                    json.dump(db_existente, f, ensure_ascii=False, indent=4)
                    
                self.finished.emit(f"Banco de dados atualizado: {count} registros verificados.")
            else:
                self.finished.emit("Falha ao conectar com servidor de mapas.")
        except Exception as e:
            self.finished.emit(f"Erro na atualização automática: {str(e)}")

# --- WORKER THREAD (Processamento em Segundo Plano) ---
class RouteWorker(QThread):
    update_progress = pyqtSignal(str)
    finished = pyqtSignal(list, list) # Retorna (rota, nao_encontrados)
    finished_equipes = pyqtSignal(dict, list) # Retorna (equipes_dict, nao_encontrados)
    error = pyqtSignal(str)
    
    def __init__(self, base_address, lista_bairros, num_equipes=0):
        super().__init__()
        self.base_address = base_address
        self.bairros_raw = lista_bairros
        self.num_equipes = num_equipes  # 0 = rota normal, >1 = divisão por equipes
        self.cache_file = CACHE_FILE
        self.cache = self.load_cache()
        self.db_data = self.load_database()

    def load_database(self):
        if os.path.exists(DB_FILE):
            try:
                with open(DB_FILE, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    # Cria índice normalizado (ignora maiúsculas e acentos)
                    normalized_db = {}
                    for k, v in data.items():
                        normalized_db[k] = v # Chave original
                        normalized_db[remover_acentos(k).lower()] = v # Chave universal
                    return normalized_db
            except: return {}
        return {}

    def load_cache(self):
        if os.path.exists(self.cache_file):
            try:
                with open(self.cache_file, 'r', encoding='utf-8') as f:
                    cache = json.load(f)
                    # Remove entradas problemáticas conhecidas (como o Rio Três Barras) para forçar nova busca
                    keys_to_remove = [k for k in cache if "tres barras" in normalizar_endereco(k)]
                    for k in keys_to_remove:
                        del cache[k]
                    return cache
            except: return {}
        return {}

    def save_cache(self):
        try:
            with open(self.cache_file, 'w', encoding='utf-8') as f:
                json.dump(self.cache, f, ensure_ascii=False, indent=4)
        except: pass
        
    def clear_cache(self):
        self.cache = {}
        if os.path.exists(self.cache_file):
            os.remove(self.cache_file)
        
    def run(self):
        try:
            # Limpa e prepara a lista
            destinos = [b.strip() for b in self.bairros_raw if b.strip()]
            if not destinos:
                self.error.emit("A lista de destinos está vazia.")
                return
            
            # Adiciona a Base Fixa no início da lista para o roteamento
            bairros = [self.base_address] + destinos

            # Configura Geocodificador
            geolocator = Nominatim(user_agent="meu_roteirizador_cuiaba_app")
            geocode = RateLimiter(geolocator.geocode, min_delay_seconds=1)
            geolocator_arcgis = ArcGIS()
            
            dados_coletados = [] # Lista de dicts: {'lat', 'lon', 'nome', 'tipo'}
            nao_encontrados = [] # Lista de bairros não encontrados
            
            self.update_progress.emit("Buscando coordenadas...")
            
            # Prepara chaves do banco para busca fuzzy (otimização)
            chaves_db = list(self.db_data.keys()) if self.db_data else []
            
            cidade_config = APP_CONFIG.get('cidade', 'Cuiabá')
            estado_config = APP_CONFIG.get('estado', 'MT')
            
            for bairro in bairros:
                # 1. Normaliza o nome (expande abreviações como 'adm', 'jd', etc.)
                bairro_clean = normalizar_endereco(bairro)
                
                # CRIA VARIAÇÕES DE NOME (Ex: "Jardim Novo Horizonte" -> tenta também "Novo Horizonte")
                # Isso resolve o problema de bairros que em Cuiabá não têm "Jardim" no nome oficial
                variacoes_nome = [bairro_clean]
                
                # Adiciona versão original (ex: "JD Brasil") caso a expansão para "Jardim" falhe
                # Isso ajuda se o mapa (OSM/ArcGIS) estiver cadastrado como "Jd." ou "JD"
                bairro_raw_clean = " ".join(bairro.replace('.', ' ').split())
                if bairro_raw_clean.lower() != bairro_clean.lower():
                    variacoes_nome.append(bairro_raw_clean)

                prefixos_ignoraveis = ["jardim ", "residencial ", "parque ", "vila ", "setor ", "loteamento "]
                b_lower = bairro_clean.lower()
                for p in prefixos_ignoraveis:
                    if b_lower.startswith(p):
                        variacoes_nome.append(bairro_clean[len(p):].strip())
                
                # Detecta se é busca por condomínio (ex: "Cond. Florais", "Condominio X")
                eh_busca_condominio = "cond" in bairro_clean.lower()
                
                lat, lon = None, None
                origem_dado = "desconhecido" # json, geopy, api
                
                # --- TENTATIVA 0: COORDENADAS DIRETAS ---
                # Se o input for coordenadas (ex: "-15.63..., -56.02...")
                if ',' in bairro:
                    try:
                        parts = bairro.split(',')
                        if len(parts) == 2:
                            c_lat = float(parts[0].strip())
                            c_lon = float(parts[1].strip())
                            if validar_coordenadas_raw(c_lat, c_lon):
                                lat, lon = c_lat, c_lon
                                origem_dado = "coordenadas"
                                self.update_progress.emit(f"[OK] Coordenadas exatas: {bairro}")
                    except: pass
                
                # Verifica se tem números (indica endereço exato vs bairro genérico)
                tem_numeros = any(char.isdigit() for char in bairro_clean)
                
                # --- TENTATIVA 0: CACHE (Memória Rápida) ---
                if not lat:
                    self.update_progress.emit(f"Verificando Cache: {bairro}...")
                    if bairro_clean in self.cache:
                        data = self.cache[bairro_clean]
                        c_lat, c_lon = data.get('lat'), data.get('lon')
                        
                        if validar_coordenadas_raw(c_lat, c_lon):
                            lat, lon = c_lat, c_lon
                            origem_dado = data.get('origem', 'geopy') # Assume geopy se antigo
                            self.update_progress.emit(f"[OK] Encontrado (Cache): {bairro}")
                        else:
                            self.update_progress.emit(f"Cache ignorado (fora da região): {bairro}")
                
                # --- TENTATIVA 1: BANCO DE DADOS LOCAL (Prioridade Máxima) ---
                if not lat and self.db_data:
                    self.update_progress.emit(f"Buscando no Banco Local: {bairro}...")
                    
                    for nome_busca in variacoes_nome:
                        if lat: break
                        
                        # Busca flexível (Tenta exato, expandido ou totalmente normalizado)
                        chave_norm = remover_acentos(nome_busca).lower()
                        cand_db = self.db_data.get(bairro) or self.db_data.get(nome_busca) or self.db_data.get(chave_norm)
                        
                        if cand_db:
                            c_lat, c_lon = cand_db.get('lat'), cand_db.get('lon')
                            if validar_coordenadas_raw(c_lat, c_lon):
                                lat, lon = c_lat, c_lon
                                origem_dado = "json"
                                self.update_progress.emit(f"[OK] Encontrado (Banco Local): {nome_busca}")

                    # --- NOVO: Busca Fuzzy (Similaridade) ---
                    # Se não achou exato, tenta achar algo parecido no banco (ex: "Bela Vista" -> "Bella Vista")
                    if not lat and chaves_db:
                        for nome_busca in variacoes_nome:
                            if lat: break
                            chave_norm = remover_acentos(nome_busca).lower()
                            # Cutoff 0.92 exige alta similaridade para evitar erros grosseiros
                            matches = difflib.get_close_matches(chave_norm, chaves_db, n=1, cutoff=0.92)
                            if matches:
                                melhor_match = matches[0]
                                cand_db = self.db_data[melhor_match]
                                c_lat, c_lon = cand_db.get('lat'), cand_db.get('lon')
                                if validar_coordenadas_raw(c_lat, c_lon):
                                    lat, lon = c_lat, c_lon
                                    origem_dado = f"json (similar: {melhor_match})"
                                    self.update_progress.emit(f"[OK] Encontrado por similaridade: {nome_busca} -> {melhor_match}")

                    if not lat:
                        self.update_progress.emit(f"Não encontrado no Banco Local.")
                        time.sleep(0.5)

                # --- TENTATIVA 2: GEOCODIFICAÇÃO (Geopy) ---
                if not lat:
                    self.update_progress.emit(f"Buscando no Geopy: {bairro}...")
                    
                    # Lista de cidades para tentar
                    cidades_alvo = [cidade_config]
                    
                    for nome_busca in variacoes_nome:
                        if lat: break
                        
                        for cidade in cidades_alvo:
                            if lat: break # Se já achou na cidade anterior, para
                            
                            queries = [
                                # Tenta busca estruturada (Mais precisa)
                                {'neighborhood': nome_busca, 'city': cidade, 'state': estado_config},
                                # Tenta busca "Bairro X"
                                f"Bairro {nome_busca}, {cidade}, {estado_config}",
                            ]
                            
                            # Busca genérica apenas se for condomínio (evita falsos positivos de ruas/lojas)
                            if eh_busca_condominio:
                                queries.append(f"{nome_busca}, {cidade}, {estado_config}")
                            
                            for q in queries:
                                try:
                                    cand = geocode(q)
                                    
                                    # Ajuste: Se busca condomínio, permite POIs e valida nome. Se não, bloqueia condomínios.
                                    if cand and validar_localizacao(cand, ignorar_pois=not eh_busca_condominio, permitir_condominio=eh_busca_condominio) and eh_bairro_de_verdade(cand, nome_busca):
                                        lat, lon = cand.latitude, cand.longitude
                                        origem_dado = f"geopy ({cidade})"
                                        self.update_progress.emit(f"[OK] Encontrado em {cidade}: {nome_busca}")
                                        break
                                except: pass
                            
                    if not lat:
                        self.update_progress.emit(f"Não encontrado no Geopy.")
                        time.sleep(0.5)

                # --- TENTATIVA 2.5: ARCGIS (Backup Gratuito e Robusto) ---
                if not lat:
                    self.update_progress.emit(f"Buscando no ArcGIS: {bairro}...")
                    
                    for nome_busca in variacoes_nome:
                        if lat: break
                        try:
                            # ArcGIS é excelente para encontrar bairros que o OSM falha
                            loc_arc = geolocator_arcgis.geocode(f"{nome_busca}, {cidade_config}, {estado_config}", timeout=5)
                            
                            # Validação extra de texto para ArcGIS (evita resultados aleatórios)
                            texto_valido = False
                            if loc_arc:
                                addr_arc = remover_acentos(loc_arc.address.lower())
                                ignorar = ['bairro', 'jardim', 'residencial', 'condominio', 'parque', 'vila', 'setor']
                                busca_parts = [p for p in remover_acentos(nome_busca.lower()).split() if len(p) > 2 and p not in ignorar]
                                
                                # Se não tiver palavras significativas (ex: "Jardim"), aceita. Se tiver, verifica presença.
                                if not busca_parts or any(part in addr_arc for part in busca_parts):
                                    texto_valido = True

                            if loc_arc and texto_valido and validar_localizacao(loc_arc, ignorar_pois=not eh_busca_condominio, permitir_condominio=eh_busca_condominio) and eh_bairro_de_verdade(loc_arc, nome_busca):
                                lat, lon = loc_arc.latitude, loc_arc.longitude
                                origem_dado = "arcgis"
                                self.update_progress.emit(f"[OK] Encontrado (ArcGIS): {nome_busca}")
                                break
                        except Exception as e:
                            print(f"Erro ArcGIS: {e}")
                    
                    if not lat:
                        self.update_progress.emit(f"Não encontrado no ArcGIS.")
                        time.sleep(0.5)

                # --- TENTATIVA 3: API DE CEP (Fallback Final) ---
                if not lat:
                    # Verifica se tem CEP na string (Regex para XXXXX-XXX ou XXXXXXXX)
                    cep_match = re.search(r'\b\d{5}-?\d{3}\b', bairro)
                    if cep_match:
                        cep_candidate = cep_match.group().replace('-', '')
                        try:
                            self.update_progress.emit(f"Buscando CEP {cep_candidate} na API...")
                            url = f"https://brasilapi.com.br/api/cep/v2/{cep_candidate}"
                            resp = requests.get(url, timeout=5)
                            if resp.status_code == 200:
                                data_api = resp.json()
                                
                                # VALIDAÇÃO DE CIDADE NA API (NOVO)
                                if data_api.get('city', '').lower() != cidade_config.lower():
                                    self.update_progress.emit(f"Ignorado: CEP de {data_api.get('city')}")
                                    continue
                                    
                                if 'location' in data_api and 'coordinates' in data_api['location']:
                                    coords = data_api['location']['coordinates']
                                    lat = float(coords.get('latitude'))
                                    lon = float(coords.get('longitude'))
                                    origem_dado = "api"
                                    self.update_progress.emit(f"[OK] Encontrado (BrasilAPI): {bairro}")
                                else:
                                    self.update_progress.emit(f"CEP sem coordenadas.")
                            else:
                                self.update_progress.emit(f"CEP não encontrado.")
                                time.sleep(0.5)
                        except Exception as e:
                            print(f"Erro API CEP: {e}")
                            self.update_progress.emit(f"Erro na API de CEP.")
                            time.sleep(0.5)

                # --- CONSOLIDAÇÃO ---
                if lat and lon:
                    zona = identificar_zona(lat, lon)
                    dados_coletados.append({
                        'lat': lat,
                        'lon': lon,
                        'nome': f"{bairro} ({zona})",
                        'tipo': origem_dado
                    })
                    
                    # Atualiza Cache
                    self.cache[bairro_clean] = {
                        'lat': lat, 
                        'lon': lon,
                        'address': f"Origem: {origem_dado}",
                        'origem': origem_dado
                    }
                else:
                    self.update_progress.emit(f"[FALHA] Não encontrado: {bairro}")
                    nao_encontrados.append(bairro)

            # Salva o cache atualizado em disco
            self.save_cache()
            
            if len(dados_coletados) < 2:
                self.error.emit("Poucos endereços encontrados para traçar rota (Mínimo 2).")
                return
            
            # --- MODO DIVISÃO POR EQUIPES ---
            if self.num_equipes > 1:
                self.update_progress.emit(f"Dividindo {len(dados_coletados)-1} serviços em {self.num_equipes} equipes...")
                base_dados = dados_coletados[0]  # Primeiro item é sempre a base
                servicos = dados_coletados[1:]   # Restante são os serviços
                
                equipes = dividir_em_equipes(servicos, self.num_equipes, base_dados)
                
                self.update_progress.emit(f"Divisão concluída: {len(equipes)} equipes criadas.")
                self.finished_equipes.emit(equipes, nao_encontrados)
            else:
                # --- MODO ROTA NORMAL ---
                self.update_progress.emit("Otimizando rota (TSP)...")
                rota_final = organizar_rota_tsp(dados_coletados)
                
                # Fecha o ciclo (Retorno à base)
                if rota_final:
                    p_retorno = rota_final[0].copy()
                    p_retorno['nome'] += " (Retorno)"
                    rota_final.append(p_retorno)
                
                # Emite a lista de dados para a UI processar e gerar o mapa
                self.finished.emit(rota_final, nao_encontrados)
            
            if nao_encontrados:
                self.update_progress.emit(f"Itens não encontrados: {', '.join(nao_encontrados)}")
            
        except Exception as e:
            self.error.emit(str(e))

# --- INTERFACE GRÁFICA (GUI) ---
class LocalizadorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Roteirizador Manutenção - BrPaving")
        self.resize(1200, 900)
        load_config() # Carrega configurações globais
        self.load_settings()
        self.rota_atual = [] # Armazena a lista de dicionários da rota
        self.col_endereco_atual = None # Armazena o nome da coluna de endereço usada
        self.nao_encontrados = [] # Armazena itens não encontrados
        self.map_worker = None  # Worker para regenerar mapa
        self.db_updater = None  # Worker de atualização
        self.equipes_resultado = {}  # Dict de equipes divididas
        self.equipe_selecionada = None  # Nome da equipe atualmente exibida
        
        # Configuração do Ícone (Basta colocar um arquivo 'icone.png' na mesma pasta)
        icon_path = os.path.join(SCRIPT_DIR, "icone.ico")
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
            
        # Verifica se é a primeira execução (sem config)
        if not os.path.exists(CONFIG_FILE):
            dlg = ConfigDialog(self)
            dlg.exec_()
        
        # Widget Central
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Layout principal (apenas para o mapa preencher tudo)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        
        # Visualizador Web (Mapa) - Fundo
        self.web_view = QWebEngineView()
        
        # Configura página personalizada para interceptar cliques
        self.web_page = CustomWebEnginePage(self.web_view)
        self.web_page.callback_add_bairro = self.abrir_adicionar_bairro
        self.web_view.setPage(self.web_page)
        
        main_layout.addWidget(self.web_view)
        self.web_view.setHtml(self.gerar_mapa_inicial())
        
        # --- OVERLAY DE LOADING (Círculo de Carregamento) ---
        self.loading_overlay = QFrame(central_widget)
        self.loading_overlay.setStyleSheet("""
            QFrame { background-color: rgba(255, 255, 255, 0.8); border-radius: 10px; }
            QLabel { color: #004aad; font-weight: bold; font-size: 16px; }
        """)
        self.loading_overlay.hide()
        load_layout = QVBoxLayout(self.loading_overlay)
        
        # Label com texto piscando ou simples
        self.lbl_loading = QLabel("Processando Rota...\nAguarde")
        self.lbl_loading.setAlignment(Qt.AlignCenter)
        load_layout.addWidget(self.lbl_loading)
        
        # Barra de progresso infinita (Marquee)
        self.loading_bar = QProgressBar()
        self.loading_bar.setRange(0, 0) # Modo indeterminado (vai e volta)
        self.loading_bar.setTextVisible(False)
        self.loading_bar.setStyleSheet("""
            QProgressBar { border: 1px solid #bbb; border-radius: 4px; background: white; height: 10px; }
            QProgressBar::chunk { background-color: #004aad; border-radius: 4px; }
        """)
        load_layout.addWidget(self.loading_bar)
        
        self.loading_overlay.setFixedSize(200, 100)
        
        # --- HUD Flutuante (Lateral) ---
        self.hud_widget = QWidget(central_widget)
        self.hud_widget.setObjectName("HudWidget")
        self.hud_widget.move(20, 20)
        self.hud_widget.setFixedWidth(360)
        self.hud_widget.setStyleSheet("""
            #HudWidget {
                background-color: rgba(255, 255, 255, 0.95); 
                border-radius: 15px; 
                border: 1px solid #e0e0e0;
                font-family: 'Segoe UI', sans-serif;
            }
            QScrollBar:vertical {
                border: none;
                background: #f0f0f0;
                width: 8px;
                margin: 0px;
                border-radius: 4px;
            }
            QScrollBar::handle:vertical {
                background: #c0c0c0;
                min-height: 20px;
                border-radius: 4px;
            }
            QScrollBar::handle:vertical:hover {
                background: #a0a0a0;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px;
            }
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
                background: none;
            }
        """)
        
        main_hud_layout = QVBoxLayout(self.hud_widget)
        main_hud_layout.setContentsMargins(10, 10, 10, 10)
        main_hud_layout.setSpacing(5)

        # --- HEADER (Título + Botão Minimizar) ---
        header_widget = QWidget()
        header_layout = QHBoxLayout(header_widget)
        header_layout.setContentsMargins(0, 0, 0, 5)
        
        # Título de Boas Vindas (Removido a pedido do usuário para mais espaço)
        
        header_layout.addStretch()
        
        # Botão Importar DB (Icone)
        self.btn_import_db = QPushButton()
        self.btn_import_db.setIcon(qta.icon('fa5s.file-import', color='#f0ad4e'))
        self.btn_import_db.setFixedSize(40, 40)
        self.btn_import_db.setCursor(Qt.PointingHandCursor)
        self.btn_import_db.setToolTip("Importar Banco de Dados")
        self.btn_import_db.setStyleSheet("QPushButton { border: none; background: transparent; border-radius: 20px; } QPushButton:hover { background-color: #e0e0e0; }")
        self.btn_import_db.clicked.connect(self.importar_banco)
        header_layout.addWidget(self.btn_import_db)
        header_layout.addStretch()

        # Botão Exportar DB (Icone)
        self.btn_export_db = QPushButton()
        self.btn_export_db.setIcon(qta.icon('fa5s.file-export', color='#17a2b8'))
        self.btn_export_db.setFixedSize(40, 40)
        self.btn_export_db.setCursor(Qt.PointingHandCursor)
        self.btn_export_db.setToolTip("Exportar Banco de Dados")
        self.btn_export_db.setStyleSheet("QPushButton { border: none; background: transparent; border-radius: 20px; } QPushButton:hover { background-color: #e0e0e0; }")
        self.btn_export_db.clicked.connect(self.exportar_banco)
        header_layout.addWidget(self.btn_export_db)
        header_layout.addStretch()
        
        # Botão Configurações (Icone)
        self.btn_config = QPushButton()
        self.btn_config.setIcon(qta.icon('fa5s.cog', color='#6c757d'))
        self.btn_config.setFixedSize(40, 40)
        self.btn_config.setCursor(Qt.PointingHandCursor)
        self.btn_config.setToolTip("Configurações (Cidade/Base)")
        self.btn_config.setStyleSheet("QPushButton { border: none; background: transparent; border-radius: 20px; } QPushButton:hover { background-color: #e0e0e0; }")
        self.btn_config.clicked.connect(self.abrir_config)
        header_layout.addWidget(self.btn_config)
        header_layout.addStretch()

        self.btn_center = QPushButton()
        self.btn_center.setIcon(qta.icon('fa5s.crosshairs', color='#004aad'))
        self.btn_center.setFixedSize(40, 40)
        self.btn_center.setCursor(Qt.PointingHandCursor)
        self.btn_center.setToolTip("Centralizar Rota")
        self.btn_center.setStyleSheet("QPushButton { border: none; background: transparent; border-radius: 20px; } QPushButton:hover { background-color: #e0e0e0; }")
        self.btn_center.clicked.connect(self.expandir_mapa)
        self.btn_center.setEnabled(False)
        header_layout.addWidget(self.btn_center)
        header_layout.addStretch()

        self.btn_toggle = QPushButton()
        self.btn_toggle.setIcon(qta.icon('fa5s.chevron-up', color='#004aad'))
        self.btn_toggle.setFixedSize(40, 40)
        self.btn_toggle.setCursor(Qt.PointingHandCursor)
        self.btn_toggle.setToolTip("Minimizar/Expandir")
        self.btn_toggle.setStyleSheet("QPushButton { border: none; background: transparent; border-radius: 20px; } QPushButton:hover { background-color: #e0e0e0; }")
        self.btn_toggle.clicked.connect(self.toggle_hud)
        header_layout.addWidget(self.btn_toggle)
        header_layout.addStretch()
        
        main_hud_layout.addWidget(header_widget)
        
        # --- BODY (Conteúdo) ---
        self.hud_body = QWidget()
        main_hud_layout.addWidget(self.hud_body)
        
        hud_layout = QVBoxLayout(self.hud_body)
        hud_layout.setContentsMargins(0, 0, 0, 0)
        hud_layout.setSpacing(10)
        
        # Base Dinâmica
        lbl_base = QLabel("<b>Base de Saída:</b>")
        lbl_base.setStyleSheet("background: transparent; border: none;")
        hud_layout.addWidget(lbl_base)
        
        self.txt_base = QLineEdit(APP_CONFIG.get('base_address', ''))
        self.txt_base.setStyleSheet("background: white; border: 1px solid #ccc; border-radius: 4px; padding: 4px;")
        hud_layout.addWidget(self.txt_base)
        
        # --- IMPORTAÇÃO EXCEL ---
        self.btn_importar = QPushButton(" Importar Excel")
        self.btn_importar.setIcon(qta.icon('fa5s.file-excel', color='white'))
        self.btn_importar.setStyleSheet("background-color: #28a745; color: white; font-weight: bold; padding: 10px; font-size: 12px; border-radius: 6px; margin-top: 5px;")
        self.btn_importar.clicked.connect(self.carregar_excel)
        hud_layout.addWidget(self.btn_importar)

        # Número de Equipes
        lbl_num_equipes = QLabel("<b>Número de Equipes (1-50):</b>")
        lbl_num_equipes.setStyleSheet("background: transparent; border: none; margin-top: 5px;")
        hud_layout.addWidget(lbl_num_equipes)

        self.spin_num_equipes = QSpinBox()
        self.spin_num_equipes.setMinimum(1)
        self.spin_num_equipes.setMaximum(50)
        self.spin_num_equipes.setValue(1)
        self.spin_num_equipes.setStyleSheet("background: white; border: 1px solid #ccc; border-radius: 4px; padding: 4px;")
        hud_layout.addWidget(self.spin_num_equipes)

        # Botão de Divisão
        self.btn_dividir = QPushButton(" Dividir e Roteirizar")
        self.btn_dividir.setIcon(qta.icon('fa5s.users', color='white'))
        self.btn_dividir.setStyleSheet("background-color: #28a745; color: white; font-weight: bold; padding: 12px; font-size: 14px; border-radius: 8px; margin-top: 5px;")
        self.btn_dividir.clicked.connect(self.iniciar_divisao)
        self.btn_dividir.setEnabled(False) # Habilita ao carregar excel
        hud_layout.addWidget(self.btn_dividir)

        # Combo Resultado das Equipes
        lbl_equipe_res = QLabel("<b>Visualizar Equipe:</b>")
        lbl_equipe_res.setStyleSheet("background: transparent; border: none; margin-top: 5px;")
        hud_layout.addWidget(lbl_equipe_res)

        self.combo_resultado_equipes = QComboBox()
        self.combo_resultado_equipes.addItem("Aguardando Divisão...")
        self.combo_resultado_equipes.setStyleSheet("background: white; border: 1px solid #ccc; border-radius: 4px; padding: 4px;")
        self.combo_resultado_equipes.setEnabled(False)
        self.combo_resultado_equipes.currentTextChanged.connect(self.selecionar_equipe_resultado)
        hud_layout.addWidget(self.combo_resultado_equipes)

        lbl_dest = QLabel("<b>Destinos Livres (Opcional):</b>")
        lbl_dest.setStyleSheet("background: transparent; border: none; margin-top: 5px;")
        hud_layout.addWidget(lbl_dest)
        
        self.txt_input = QTextEdit()
        self.txt_input.setPlaceholderText("Ex: Jardim das Américas, Pedra 90...")
        self.txt_input.setMaximumHeight(60) # Entrada menor
        self.txt_input.setStyleSheet("background: white; border: 1px solid #ccc; border-radius: 4px;")
        hud_layout.addWidget(self.txt_input)
        
        self.btn_gerar = QPushButton(" Roteirizar Apenas Livres")
        self.btn_gerar.setIcon(qta.icon('fa5s.route', color='white'))
        self.btn_gerar.setStyleSheet("background-color: #004aad; color: white; font-weight: bold; padding: 12px; font-size: 14px; border-radius: 8px;")
        self.btn_gerar.clicked.connect(self.iniciar_processamento)
        self.btn_gerar.setEnabled(False)
        hud_layout.addWidget(self.btn_gerar)
        
        # Botão Limpar Cache
        self.btn_limpar = QPushButton(" Limpar Cache")
        self.btn_limpar.setIcon(qta.icon('fa5s.trash-alt', color='white'))
        self.btn_limpar.setStyleSheet("background-color: #d9534f; color: white; font-weight: bold; padding: 5px; font-size: 12px; border-radius: 5px; margin-top: 5px;")
        self.btn_limpar.clicked.connect(self.limpar_cache)
        hud_layout.addWidget(self.btn_limpar)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        self.progress_bar.hide()
        self.progress_bar.setStyleSheet("background: transparent; border: none;")
        hud_layout.addWidget(self.progress_bar)
        

        # Área de Log (Substitui o Label simples)
        self.txt_log = QTextEdit()
        self.txt_log.setReadOnly(True)
        self.txt_log.setPlaceholderText("Log de processamento...")
        self.txt_log.setStyleSheet("background: #fff; border: 1px solid #eee; border-radius: 4px; font-size: 10px; color: #666;")
        self.txt_log.setMaximumHeight(60)
        hud_layout.addWidget(self.txt_log)
        
        hud_layout.addStretch()
        
        # Assinatura do Desenvolvedor
        lbl_assinatura = QLabel("Desenvolvido por - João Eduardo")
        lbl_assinatura.setStyleSheet("color: #999; font-size: 10px; font-style: italic; margin-top: 5px;")
        lbl_assinatura.setAlignment(Qt.AlignCenter)
        hud_layout.addWidget(lbl_assinatura)
        
        # Garante que o HUD fique sobre o mapa
        self.hud_widget.raise_()
        self.hud_widget.adjustSize()
        
        # --- WIDGET DA LISTA (LADO DIREITO) ---
        # Restaura a "Aba da Lista" no lado direito, mas agora interativa
        self.lista_widget = QWidget(central_widget)
        self.lista_widget.setObjectName("ListaWidget")
        self.lista_widget.setFixedWidth(420) # Aumentado para evitar cortes nos botões
        self.lista_widget.setStyleSheet("""
            #ListaWidget {
                background-color: rgba(255, 255, 255, 0.95); 
                border-radius: 15px; 
                border: 1px solid #e0e0e0;
                box-shadow: -5px 5px 15px rgba(0,0,0,0.1);
            }
        """)
        
        lista_layout = QVBoxLayout(self.lista_widget)
        lista_layout.setContentsMargins(15, 15, 15, 15)
        lista_layout.setSpacing(10)
        
        lbl_ordem = QLabel("<b>Ordem da Rota</b> (Arraste para mudar)")
        lbl_ordem.setStyleSheet("background: transparent; border: none; color: #333;")
        lista_layout.addWidget(lbl_ordem)

        self.list_rota = QListWidget()
        self.list_rota.setDragDropMode(QAbstractItemView.InternalMove) # Permite arrastar
        self.list_rota.setWordWrap(True)
        self.list_rota.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.list_rota.setStyleSheet("""
            QListWidget {
                background: #fff; 
                border: 1px solid #eee; 
                border-radius: 8px; 
                font-size: 13px; 
                color: #444;
                padding: 5px;
            }
            QListWidget::item {
                padding: 8px;
                border-bottom: 1px solid #f0f0f0;
            }
            QListWidget::item:selected {
                background-color: #e3f2fd;
                color: #004aad;
                border-radius: 4px;
            }
        """)
        # Conecta o sinal de mudança de ordem
        self.list_rota.model().rowsMoved.connect(self.recalcular_rota_manual)
        # Conecta o clique para voar até o local
        self.list_rota.itemClicked.connect(self.focar_ponto)
        lista_layout.addWidget(self.list_rota)
        
        # Layout para botões de ação da rota
        btns_rota_layout = QHBoxLayout()
        btns_rota_layout.addStretch() # Centraliza os ícones

        # Botão Inverter Rota (Agora aqui na lista)
        self.btn_invert = QPushButton(" Inverter")
        # Botão Inverter Rota
        self.btn_invert = QPushButton()
        self.btn_invert.setIcon(qta.icon('fa5s.exchange-alt', color='white'))
        self.btn_invert.setStyleSheet("background-color: #f0ad4e; color: white; font-weight: bold; padding: 10px 5px; font-size: 11px; border-radius: 8px;")
        self.btn_invert.setFixedSize(36, 36)
        self.btn_invert.setToolTip("Inverter Ordem da Rota")
        self.btn_invert.setStyleSheet("background-color: #f0ad4e; border-radius: 18px;")
        self.btn_invert.clicked.connect(self.inverter_rota)
        self.btn_invert.setEnabled(False)
        btns_rota_layout.addWidget(self.btn_invert)
        
        # Botão Copiar Link (Movido para cá)
        self.btn_copy = QPushButton(" Copiar Link")
        # Botão Copiar Link
        self.btn_copy = QPushButton()
        self.btn_copy.setIcon(qta.icon('fa5s.link', color='white'))
        self.btn_copy.setStyleSheet("background-color: #00c853; color: white; font-weight: bold; padding: 10px 5px; font-size: 11px; border-radius: 8px;")
        self.btn_copy.setFixedSize(36, 36)
        self.btn_copy.setToolTip("Copiar Link do Google Maps")
        self.btn_copy.setStyleSheet("background-color: #00c853; border-radius: 18px;")
        self.btn_copy.clicked.connect(self.copiar_link)
        self.btn_copy.setEnabled(False)
        btns_rota_layout.addWidget(self.btn_copy)
        
        # Botão Copiar Lista (Ordem em Texto)
        self.btn_copy_list = QPushButton(" Copiar Lista")
        # Botão Copiar Lista
        self.btn_copy_list = QPushButton()
        self.btn_copy_list.setIcon(qta.icon('fa5s.list-ol', color='white'))
        self.btn_copy_list.setStyleSheet("background-color: #6c757d; color: white; font-weight: bold; padding: 10px 5px; font-size: 11px; border-radius: 8px;")
        self.btn_copy_list.setFixedSize(36, 36)
        self.btn_copy_list.setToolTip("Copiar Lista de Endereços")
        self.btn_copy_list.setStyleSheet("background-color: #6c757d; border-radius: 18px;")
        self.btn_copy_list.clicked.connect(self.copiar_lista_rota)
        self.btn_copy_list.setEnabled(False)
        btns_rota_layout.addWidget(self.btn_copy_list)
        
        # Botão Exportar Excel (Novo)
        self.btn_export = QPushButton(" Exportar Excel")
        # Botão Exportar Excel
        self.btn_export = QPushButton()
        self.btn_export.setIcon(qta.icon('fa5s.file-export', color='white'))
        self.btn_export.setStyleSheet("background-color: #17a2b8; color: white; font-weight: bold; padding: 10px 5px; font-size: 11px; border-radius: 8px;")
        self.btn_export.setFixedSize(36, 36)
        self.btn_export.setToolTip("Exportar para Excel")
        self.btn_export.setStyleSheet("background-color: #17a2b8; border-radius: 18px;")
        self.btn_export.clicked.connect(self.exportar_excel)
        self.btn_export.setEnabled(False)
        btns_rota_layout.addWidget(self.btn_export)
        
        btns_rota_layout.addStretch() # Centraliza os ícones

        lista_layout.addLayout(btns_rota_layout)
        
        self.lista_widget.hide() # Começa escondido até ter rota
        
        # --- INICIA ATUALIZAÇÃO AUTOMÁTICA ---
        self.iniciar_atualizacao_banco()

    def gerar_mapa_inicial(self):
        # Gera um mapa limpo focado na cidade configurada
        center = APP_CONFIG.get('center', [-15.5989, -56.0949])
        m = folium.Map(location=center, zoom_start=13, control_scale=True)
        data = io.BytesIO()
        m.save(data, close_file=False)
        return data.getvalue().decode()
        
    def abrir_config(self):
        dlg = ConfigDialog(self)
        if dlg.exec_():
            # Recarrega base e atualiza mapa inicial
            self.txt_base.setText(APP_CONFIG.get('base_address', ''))
            self.web_view.setHtml(self.gerar_mapa_inicial())
            self.iniciar_atualizacao_banco() # Re-baixa dados da nova cidade

    def iniciar_atualizacao_banco(self):
        """Roda a atualização do OSM em background sem travar a UI."""
        self.txt_log.append("Verificando atualizações de mapas...")
        self.db_updater = DatabaseUpdater()
        self.db_updater.finished.connect(lambda msg: self.txt_log.append(f"INFO: {msg}"))
        self.db_updater.start()

    def toggle_hud(self):
        if self.hud_body.isVisible():
            self.hud_body.hide()
            self.btn_toggle.setIcon(qta.icon('fa5s.chevron-down', color='#004aad'))
        else:
            self.hud_body.show()
            self.btn_toggle.setIcon(qta.icon('fa5s.chevron-up', color='#004aad'))
    def resizeEvent(self, event):
        # Mantém o widget da lista ancorado à direita
        if hasattr(self, 'lista_widget'):
            self.lista_widget.move(self.width() - 440, 20) # Ajustado para nova largura (420 + 20 margem)
            self.lista_widget.setFixedHeight(self.height() - 40)
        # Centraliza o loading
        self.loading_overlay.move((self.width() - 200) // 2, (self.height() - 100) // 2)
        super().resizeEvent(event)

    def iniciar_divisao(self):
        """Inicia o processo de divisão de serviços por equipes."""
        if not hasattr(self, 'df_dados') or self.df_dados is None:
            QMessageBox.warning(self, "Aviso", "Por favor, carregue uma planilha Excel primeiro.")
            return
            
        base = self.txt_base.text()
        if not base.strip():
            QMessageBox.warning(self, "Aviso", "Por favor, defina a Base de Saída.")
            return
            
        APP_CONFIG['base_address'] = base
        save_config()
        
        num_equipes = self.spin_num_equipes.value()
        
        # Pega todos os bairros listados na caixa de texto (que carregou do excel)
        texto = self.txt_input.toPlainText()
        lista_bairros = texto.split(',') if texto.strip() else []
        
        if not lista_bairros:
            QMessageBox.warning(self, "Aviso", "Nenhum endereço encontrado para roteirizar.")
            return
            
        self.set_interface_busy(True)
        self.progress_bar.setRange(0, 0) # Modo indeterminado
        self.progress_bar.show()
        self.txt_log.clear()
        self.txt_log.append(f"Iniciando divisão em {num_equipes} equipes...")
        
        self.worker = RouteWorker(base, lista_bairros, num_equipes=num_equipes)
        self.worker.update_progress.connect(self.atualizar_status)
        self.worker.finished_equipes.connect(self.receber_dados_divisao)
        self.worker.error.connect(self.exibir_erro)
        self.worker.start()

    def receber_dados_divisao(self, equipes_dict, nao_encontrados):
        """Recebe o dicionário de equipes roteirizadas e atualiza a UI."""
        self.equipes_resultado = equipes_dict
        self.nao_encontrados = nao_encontrados
        
        self.combo_resultado_equipes.blockSignals(True)
        self.combo_resultado_equipes.clear()
        
        if equipes_dict:
            for nome_equipe in sorted(equipes_dict.keys()):
                self.combo_resultado_equipes.addItem(nome_equipe)
                
            self.combo_resultado_equipes.setEnabled(True)
            self.equipe_selecionada = self.combo_resultado_equipes.currentText()
            self.rota_atual = self.equipes_resultado[self.equipe_selecionada]
            self.txt_log.append(f"<b>Divisão concluída! Visualizando {self.equipe_selecionada}.</b>")
            self.atualizar_interface_rota()
        else:
            self.combo_resultado_equipes.addItem("Nenhuma rota gerada.")
            self.combo_resultado_equipes.setEnabled(False)
            self.set_interface_busy(False)
            
        self.combo_resultado_equipes.blockSignals(False)
        self.progress_bar.hide()
        self.lista_widget.show()

    def selecionar_equipe_resultado(self, equipe_nome):
        """Troca a visualização da rota para a equipe selecionada no combo."""
        if equipe_nome in self.equipes_resultado:
            self.equipe_selecionada = equipe_nome
            self.rota_atual = self.equipes_resultado[equipe_nome]
            self.txt_log.append(f"Visualizando {equipe_nome}...")
            self.atualizar_interface_rota()

    def iniciar_processamento(self):
        """Modo normal sem divisão de equipes."""
        texto = self.txt_input.toPlainText()
        base = self.txt_base.text()
        if not texto.strip():
            QMessageBox.warning(self, "Aviso", "Por favor, insira pelo menos um destino.")
            return
        if not base.strip():
            QMessageBox.warning(self, "Aviso", "Por favor, defina a Base de Saída.")
            return
            
        # Limpa dados de divisão se existir
        self.equipes_resultado.clear()
        self.combo_resultado_equipes.blockSignals(True)
        self.combo_resultado_equipes.clear()
        self.combo_resultado_equipes.addItem("Aguardando Divisão...")
        self.combo_resultado_equipes.setEnabled(False)
        self.combo_resultado_equipes.blockSignals(False)
            
        APP_CONFIG['base_address'] = base
        save_config()
            
        lista_bairros = texto.split(',')
        
        self.set_interface_busy(True)
        self.progress_bar.setRange(0, 0) # Modo indeterminado
        self.progress_bar.show()
        self.txt_log.clear()
        self.txt_log.append("Iniciando processamento de rota única...")
        
        self.worker = RouteWorker(base, lista_bairros, num_equipes=0)
        self.worker.update_progress.connect(self.atualizar_status)
        self.worker.finished.connect(self.receber_dados_rota)
        self.worker.error.connect(self.exibir_erro)
        self.worker.start()
        
    def limpar_cache(self):
        reply = QMessageBox.question(self, 'Limpar Cache', 
                                     "Tem certeza? Isso apagará todos os endereços salvos e fará a busca demorar mais na próxima vez.",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            if hasattr(self, 'worker'): self.worker.clear_cache()
            else: RouteWorker("", []).clear_cache() # Instância temporária para limpar
            QMessageBox.information(self, "Sucesso", "Cache limpo com sucesso!")

    def atualizar_status(self, msg):
        self.txt_log.append(msg)
        self.txt_log.verticalScrollBar().setValue(self.txt_log.verticalScrollBar().maximum())
        
    def copiar_link(self):
        clipboard = QApplication.clipboard()
        clipboard.setText(self.url_maps)
        QMessageBox.information(self, "Sucesso", "Link da rota copiado para a área de transferência!")
        # Feedback visual (Balãozinho)
        point = self.btn_copy.mapToGlobal(self.btn_copy.rect().topLeft())
        point.setX(point.x() + self.btn_copy.width() // 2)
        point.setY(point.y() - 10)
        QToolTip.showText(point, "Copiado!", self.btn_copy)

    def copiar_lista_rota(self):
        """Copia a ordem da rota em formato de texto simples."""
        if not self.rota_atual: return
        texto = ""
        for i, ponto in enumerate(self.rota_atual):
            nome_limpo = re.sub(r" \(Zona .*\)", "", ponto['nome'])
            nome_limpo = nome_limpo.replace(" (Retorno)", "")
            texto += f"{i+1}. {nome_limpo}\n"
        
        clipboard = QApplication.clipboard()
        clipboard.setText(texto)
        QMessageBox.information(self, "Sucesso", "Lista de endereços copiada!")
        # Feedback visual (Balãozinho)
        point = self.btn_copy_list.mapToGlobal(self.btn_copy_list.rect().topLeft())
        point.setX(point.x() + self.btn_copy_list.width() // 2)
        point.setY(point.y() - 10)
        QToolTip.showText(point, "Copiado!", self.btn_copy_list)

    def expandir_mapa(self):
        self.web_view.page().runJavaScript("fitRouteBounds();")

    def set_interface_busy(self, busy):
        """Bloqueia/Desbloqueia botões e mostra loading durante operações pesadas."""
        self.mostrar_loading(busy)
        enable = not busy
        self.btn_gerar.setEnabled(enable)
        if hasattr(self, 'btn_dividir'):
            self.btn_dividir.setEnabled(enable and hasattr(self, 'df_dados') and self.df_dados is not None)
        self.btn_invert.setEnabled(enable)
        self.btn_copy.setEnabled(enable)
        self.btn_copy_list.setEnabled(enable)
        self.btn_export.setEnabled(enable)
        self.btn_center.setEnabled(enable)
        if hasattr(self, 'combo_resultado_equipes'):
            self.combo_resultado_equipes.setEnabled(enable and len(self.equipes_resultado) > 0)
        if busy:
            self.lbl_loading.setText("Gerando Trajeto...\nAguarde")

    def receber_dados_rota(self, lista_rota, nao_encontrados):
        """Recebe a lista bruta do Worker e renderiza."""
        self.rota_atual = lista_rota
        self.nao_encontrados = nao_encontrados
        self.atualizar_interface_rota()
        self.progress_bar.hide()
        self.lista_widget.show()
        self.txt_log.append("<b>Rota gerada com sucesso!</b>")

    def inverter_rota(self):
        """Inverte a ordem dos destinos (mantendo Base no início e fim)."""
        if len(self.rota_atual) > 3: # Base, A, B, Base (minimo)
            # Fatia: [Base] + [Reverso do meio] + [Base]
            meio = self.rota_atual[1:-1]
            self.rota_atual = [self.rota_atual[0]] + meio[::-1] + [self.rota_atual[-1]]
            
            # Se faz parte de uma divisão de equipes, salva no dict
            if self.equipe_selecionada and self.equipe_selecionada in self.equipes_resultado:
                self.equipes_resultado[self.equipe_selecionada] = self.rota_atual
                
            self.atualizar_interface_rota()
            self.txt_log.append("Rota invertida.")

    def focar_ponto(self, item):
        """Voa até o ponto clicado na lista."""
        dados = item.data(Qt.UserRole)
        if dados:
            lat, lon = dados['lat'], dados['lon']
            self.web_view.page().runJavaScript(f"focusOnMarker({lat}, {lon});")

    def recalcular_rota_manual(self):
        """Chamado quando o usuário arrasta itens na lista."""
        # Reconstrói self.rota_atual baseado na ordem visual da lista
        nova_rota = []
        for i in range(self.list_rota.count()):
            item = self.list_rota.item(i)
            dados = item.data(Qt.UserRole)
            if dados: # Ignora itens que não são rota (ex: não encontrados)
                nova_rota.append(dados)
        
        self.rota_atual = nova_rota
        # Atualiza apenas o mapa e link (não precisa recriar a lista pois já está lá)
        # Mas para atualizar as distâncias (km), é bom recriar os textos
        self.atualizar_interface_rota(apenas_mapa=False)

    def mostrar_loading(self, show):
        if show:
            self.loading_overlay.show()
            self.loading_overlay.raise_()
        else:
            self.loading_overlay.hide()

    def finalizar_atualizacao_mapa(self, html_content):
        """Chamado quando o MapGenWorker termina."""
        self.web_view.setHtml(html_content)
        self.set_interface_busy(False)

    def atualizar_interface_rota(self, apenas_mapa=False):
        """Atualiza Mapa, Link e Lista com base em self.rota_atual."""
        if not self.rota_atual: return

        # Bloqueia interface enquanto gera o mapa (OSRM)
        self.set_interface_busy(True)

        # 1. Gera HTML do Mapa (Em Thread para não travar)
        self.map_worker = MapGenWorker(self, self.rota_atual)
        self.map_worker.finished.connect(self.finalizar_atualizacao_mapa)
        self.map_worker.start()
        
        # 2. Gera Link Google Maps
        str_coords = "/".join([f"{p['lat']},{p['lon']}" for p in self.rota_atual])
        self.url_maps = f"https://www.google.com/maps/dir/{str_coords}"
        
        # 3. Atualiza Lista Visual (com distâncias)
        if not apenas_mapa:
            self.list_rota.clear()
            for i, ponto in enumerate(self.rota_atual):
                # Calcula distância do anterior
                dist_txt = ""
                if i > 0:
                    p_ant = self.rota_atual[i-1]
                    dist = calcular_distancia((p_ant['lat'], p_ant['lon']), (ponto['lat'], ponto['lon']))
                    dist_txt = f" (+{dist:.1f} km)"
                
                nome_display = f"{i+1}. {ponto['nome']}{dist_txt}"
                
                item = QListWidgetItem()
                item.setData(Qt.UserRole, ponto) # Guarda os dados ocultos no item
                
                # Widget container para o item (Texto + Botão Editar)
                widget = QWidget()
                layout = QHBoxLayout(widget)
                layout.setContentsMargins(5, 5, 5, 5)
                layout.setSpacing(5)
                
                lbl = QLabel(nome_display)
                lbl.setWordWrap(True)
                lbl.setFixedWidth(300) # Reduzido levemente para garantir margem lateral
                layout.addWidget(lbl)
                
                # Botão de Edição (Lápis)
                btn_edit = QPushButton()
                btn_edit.setIcon(qta.icon('fa5s.pen', color='#888'))
                btn_edit.setFixedSize(24, 24)
                btn_edit.setCursor(Qt.PointingHandCursor)
                btn_edit.setToolTip("Editar coordenadas deste local")
                btn_edit.setStyleSheet("background: transparent; border: none;")
                btn_edit.clicked.connect(lambda checked, p=ponto: self.abrir_editar_bairro(p))
                layout.addWidget(btn_edit)
                
                # Adiciona margem extra na altura para evitar corte do texto (descendentes g, j, p, q, y)
                sz = widget.sizeHint()
                sz.setHeight(sz.height() + 10)
                item.setSizeHint(sz)
                
                # Estiliza Base e Retorno para destacar
                if i == 0 or i == len(self.rota_atual) - 1:
                    widget.setStyleSheet("background-color: #f0f0f0; border-radius: 4px;")
                    item.setFlags(item.flags() & ~Qt.ItemIsDragEnabled) # Impede arrastar a base (opcional)
                
                self.list_rota.addItem(item)
                self.list_rota.setItemWidget(item, widget)
            
            # Adiciona itens não encontrados ao final da lista
            if hasattr(self, 'nao_encontrados') and self.nao_encontrados:
                for nf in self.nao_encontrados:
                    item_nf = QListWidgetItem()
                    # Define flags para não ser arrastável nem selecionável da mesma forma que a rota
                    item_nf.setFlags(item_nf.flags() & ~Qt.ItemIsDragEnabled & ~Qt.ItemIsSelectable)
                    
                    widget_nf = QWidget()
                    layout_nf = QHBoxLayout(widget_nf)
                    layout_nf.setContentsMargins(5, 5, 5, 5)
                    
                    lbl_nf = QLabel(f"{nf} (Não encontrado)")
                    lbl_nf.setStyleSheet("color: #d9534f; font-weight: bold; font-size: 12px;")
                    lbl_nf.setWordWrap(True)
                    lbl_nf.setFixedWidth(300)
                    
                    btn_add = QPushButton("+")
                    btn_add.setFixedSize(24, 24)
                    btn_add.setCursor(Qt.PointingHandCursor)
                    btn_add.setToolTip(f"Adicionar '{nf}' ao banco de dados")
                    btn_add.setStyleSheet("background-color: #004aad; color: white; border-radius: 12px; font-weight: bold; border: none;")
                    btn_add.clicked.connect(lambda checked, nome=nf: self.abrir_adicionar_bairro(nome))
                    
                    layout_nf.addWidget(lbl_nf)
                    layout_nf.addStretch()
                    layout_nf.addWidget(btn_add)
                    
                    sz_nf = widget_nf.sizeHint()
                    sz_nf.setHeight(sz_nf.height() + 10)
                    item_nf.setSizeHint(sz_nf)
                    self.list_rota.addItem(item_nf)
                    self.list_rota.setItemWidget(item_nf, widget_nf)

    def gerar_mapa_html(self, rota):
        """Gera o código HTML do mapa Folium."""
        if not rota: return ""
        
        # Cria o mapa centrado no primeiro ponto
        mapa = folium.Map(location=[rota[0]['lat'], rota[0]['lon']], zoom_start=13)
        
        # Script JS para ajuste de zoom
        lats = [p['lat'] for p in rota]
        lons = [p['lon'] for p in rota]
        min_lat, max_lat = min(lats), max(lats)
        min_lon, max_lon = min(lons), max(lons)
        
        mapa.get_root().html.add_child(folium.Element(f"""
            <script>
                var route_bounds = [[{min_lat}, {min_lon}], [{max_lat}, {max_lon}]];
                function fitRouteBounds() {{
                    var map_instance = null;
                    for (var key in window) {{
                        if (key.startsWith('map_') && window[key].fitBounds) {{
                            map_instance = window[key];
                            break;
                        }}
                    }}
                    if (map_instance) {{
                        map_instance.fitBounds(route_bounds, {{padding: [50, 50]}});
                    }}
                }}
                
                function focusOnMarker(lat, lon) {{
                    var map_instance = null;
                    for (var key in window) {{
                        if (key.startsWith('map_') && window[key].flyTo) {{
                            map_instance = window[key];
                            break;
                        }}
                    }}
                    if (map_instance) {{
                        map_instance.flyTo([lat, lon], 16, {{duration: 1.5}});
                    }}
                }}
            </script>
        """))

        # Eixos Visuais (Cruz no centro da cidade)
        centro_lat, centro_lon = APP_CONFIG.get('center', [-15.5989, -56.0949])
        folium.PolyLine([[centro_lat - 0.15, centro_lon], [centro_lat + 0.15, centro_lon]], 
                        color="black", weight=2, opacity=0.4, dash_array="10, 10").add_to(mapa)
        folium.PolyLine([[centro_lat, centro_lon - 0.15], [centro_lat, centro_lon + 0.15]], 
                        color="black", weight=2, opacity=0.4, dash_array="10, 10").add_to(mapa)

        # Cores dos Pinos
        cores_hex = {'green': '#2aad27', 'blue': '#2a81cb', 'orange': '#cb8427', 'red': '#cb2b3e', 'gray': '#7b7b7b'}

        # Adiciona Marcadores
        for i, ponto in enumerate(rota):
            # Ignora o último ponto se for igual ao primeiro (Retorno), para não sobrepor ícone
            if i == len(rota) - 1 and i > 0:
                continue
                
            coord = [ponto['lat'], ponto['lon']]
            origem = ponto.get('tipo', 'desconhecido')
            
            if origem == 'json': cor_pino = 'green'
            elif origem == 'geopy': cor_pino = 'blue'
            elif origem == 'arcgis': cor_pino = 'orange'
            elif origem == 'api': cor_pino = 'red'
            else: cor_pino = 'gray'
            
            texto_popup = f"""
            <div style="font-family: sans-serif; min-width: 150px;">
                <b>{i+1}. {ponto['nome']}</b><br>
                <span style="color: #666; font-size: 11px;">Fonte: {origem}</span>
            </div>
            """
            
            cor_hex = cores_hex.get(cor_pino, '#7b7b7b')
            icon_html = f"""
                <div style="background-color: {cor_hex}; color: white; border-radius: 50%; width: 30px; height: 30px; 
                display: flex; justify-content: center; align-items: center; font-weight: bold; border: 2px solid white; 
                box-shadow: 2px 2px 5px rgba(0,0,0,0.4); font-family: Arial;">{i+1}</div>
            """

            folium.Marker(
                location=coord,
                popup=texto_popup,
                icon=folium.DivIcon(html=icon_html, icon_size=(30,30), icon_anchor=(15,15))
            ).add_to(mapa)

        # Desenha Linhas da Rota (Linhas retas sólidas - OSRM não confiável para esta região)
        cores_rota = ['red', 'blue', 'green', 'purple', 'orange', 'darkred', 'cadetblue', 'darkgreen', 'darkblue', 'black']
        
        for i in range(len(rota) - 1):
            p1 = [rota[i]['lat'], rota[i]['lon']]
            p2 = [rota[i+1]['lat'], rota[i+1]['lon']]
            cor_atual = cores_rota[i % len(cores_rota)]
            
            folium.PolyLine([p1, p2], color=cor_atual, weight=4, opacity=0.9).add_to(mapa)

        # Ajusta o zoom para mostrar toda a rota
        mapa.fit_bounds([[min_lat, min_lon], [max_lat, max_lon]])

        data = io.BytesIO()
        mapa.save(data, close_file=False)
        return data.getvalue().decode()
        
    def exibir_erro(self, erro_msg):
        QMessageBox.critical(self, "Erro", f"Ocorreu um erro: {erro_msg}")
        self.txt_log.append(f"<font color='red'>Erro: {erro_msg}</font>")
        self.progress_bar.hide()
        self.mostrar_loading(False)
        self.btn_gerar.setEnabled(True)

    def exportar_banco(self):
        """Exporta o banco de dados atual para um arquivo JSON para compartilhamento."""
        if not os.path.exists(DB_FILE):
            QMessageBox.warning(self, "Aviso", "Banco de dados vazio ou inexistente.")
            return
            
        fname, _ = QFileDialog.getSaveFileName(self, "Exportar Banco de Dados", "banco_bairros_compartilhar.json", "JSON Files (*.json)")
        if not fname: return
        
        try:
            shutil.copy2(DB_FILE, fname)
            QMessageBox.information(self, "Sucesso", f"Banco de dados exportado com sucesso!\nLocal: {fname}")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao exportar: {e}")

    def importar_banco(self):
        """Importa e mescla um banco de dados externo, evitando duplicatas."""
        fname, _ = QFileDialog.getOpenFileName(self, "Importar Banco de Dados", "", "JSON Files (*.json)")
        if not fname: return
        
        try:
            with open(fname, 'r', encoding='utf-8') as f:
                dados_importados = json.load(f)
            
            dados_locais = {}
            if os.path.exists(DB_FILE):
                with open(DB_FILE, 'r', encoding='utf-8') as f:
                    dados_locais = json.load(f)
            
            # Mapa de normalização para evitar duplicatas (ex: "Jardim Itália" vs "Jardim Italia")
            # Cria um dicionário { "jardim italia": "Jardim Itália" } com as chaves existentes
            mapa_norm = {remover_acentos(k).lower(): k for k in dados_locais}
            
            novos = 0
            atualizados = 0
            
            for nome, dados in dados_importados.items():
                chave_norm = remover_acentos(nome).lower()
                
                if chave_norm not in mapa_norm:
                    # Caso 1: Bairro totalmente novo
                    dados_locais[nome] = dados
                    mapa_norm[chave_norm] = nome # Atualiza o mapa temporário
                    novos += 1
                else:
                    # Caso 2: Bairro já existe (ou variação dele). Verifica conflito.
                    chave_existente = mapa_norm[chave_norm]
                    local_tipo = dados_locais[chave_existente].get('tipo', '')
                    importado_tipo = dados.get('tipo', '')
                    
                    # Regra de Ouro: Se o importado for MANUAL e o local NÃO for manual (é automático/OSM),
                    # então a versão do parceiro é melhor (foi corrigida). Atualizamos.
                    if importado_tipo == 'manual' and local_tipo != 'manual':
                        dados_locais[chave_existente] = dados # Atualiza os dados mantendo a chave local
                        atualizados += 1
            
            if novos == 0 and atualizados == 0:
                QMessageBox.information(self, "Importação", "Nenhum dado novo ou relevante encontrado.\nSeu banco já contém esses bairros.")
                return

            # Salva o Banco Atualizado
            with open(DB_FILE, 'w', encoding='utf-8') as f:
                json.dump(dados_locais, f, ensure_ascii=False, indent=4)
                
            # Limpa o cache dos itens afetados para garantir que o roteador use os novos dados
            if os.path.exists(CACHE_FILE):
                try:
                    with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                        cache = json.load(f)
                    
                    for nome in dados_importados:
                        c_norm = normalizar_endereco(nome)
                        if c_norm in cache: del cache[c_norm]
                        if nome in cache: del cache[nome]
                        
                    with open(CACHE_FILE, 'w', encoding='utf-8') as f:
                        json.dump(cache, f, ensure_ascii=False, indent=4)
                except: pass

            QMessageBox.information(self, "Sucesso", f"Importação concluída!\n\n+ {novos} novos bairros adicionados.\n+ {atualizados} bairros atualizados (melhorados).")
            self.txt_log.append(f"Importação: +{novos} novos, +{atualizados} atualizados.")

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao importar: {e}")

    def exportar_excel(self):
        """Exporta a rota atual para um arquivo Excel ordenado, incluindo equipes."""
        if not hasattr(self, 'df_dados') or self.df_dados is None or not self.col_endereco_atual:
            QMessageBox.warning(self, "Aviso", "Planilha base não encontrada ou coluna de endereço ausente.")
            return

        if not self.equipes_resultado and not self.rota_atual:
            QMessageBox.warning(self, "Aviso", "Gere ou divida uma rota primeiro.")
            return

        fname, _ = QFileDialog.getSaveFileName(self, "Salvar Excel Ordenado", "Rota_Equipes.xlsx", "Excel Files (*.xlsx)")
        if not fname: return

        try:
            df_export = self.df_dados.copy()
            
            # Garante colunas de Controle
            col_exec = "Ordem de Execução"
            col_eq = "Equipe Designada"
            if col_exec not in df_export.columns:
                df_export.insert(0, col_exec, "")
            if col_eq not in df_export.columns:
                df_export.insert(0, col_eq, "")

            # Mapa global de {endereco_limpo: (equipe, ordem)}
            mapa_end = {}
            
            if self.equipes_resultado:
                # Múltiplas equipes
                for equipe_nome, rota in self.equipes_resultado.items():
                    for i, ponto in enumerate(rota):
                        end_clean = re.sub(r" \(Zona \w+\)$", "", ponto['nome']).replace(" (Retorno)", "").lower().strip()
                        mapa_end[end_clean] = (equipe_nome, i + 1)
            else:
                # Apenas rota simples
                for i, ponto in enumerate(self.rota_atual):
                    end_clean = re.sub(r" \(Zona \w+\)$", "", ponto['nome']).replace(" (Retorno)", "").lower().strip()
                    mapa_end[end_clean] = ("Rota Única", i + 1)

            def get_equipe(end):
                end_clean = str(end).strip().lower()
                return mapa_end.get(end_clean, ("", 9999))[0]
                
            def get_ordem(end):
                end_clean = str(end).strip().lower()
                return mapa_end.get(end_clean, ("", 9999))[1]

            df_export[col_eq] = df_export[self.col_endereco_atual].apply(get_equipe)
            df_export[col_exec] = df_export[self.col_endereco_atual].apply(get_ordem)
            
            # Ordena por equipe e depois por ordem
            df_export = df_export.sort_values(by=[col_eq, col_exec])
            
            # --- Renumerar sequencialmente por equipe (1, 2, 3, 4...) ---
            # Evita números repetidos quando o mesmo bairro aparece mais de uma vez
            df_export[col_exec] = df_export.groupby(col_eq).cumcount() + 1
            
            # --- Filtrar colunas essenciais ---
            # Tenta encontrar a coluna de OS
            col_os = next((c for c in df_export.columns if c.lower().strip() in ['os', 'o.s', 'ordem', 'ordem de serviço']), None)
            
            # Tenta encontrar a coluna de Endereço
            col_endereco = next((c for c in df_export.columns if c.lower().strip() in ['endereço', 'endereco', 'endereço da os', 'endereco da os', 'logradouro', 'rua']), None)
            
            cols_to_keep = [col_eq, col_exec]
            if col_os:
                cols_to_keep.append(col_os)
            if col_endereco and col_endereco != self.col_endereco_atual:
                cols_to_keep.append(col_endereco)
            cols_to_keep.append(self.col_endereco_atual)
            
            # Filtra o DataFrame apenas para as colunas selecionadas
            df_export = df_export[cols_to_keep]
            
            df_export.to_excel(fname, index=False)
            
            # --- ESTILIZAÇÃO DO EXCEL ---
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            wb = openpyxl.load_workbook(fname)
            ws = wb.active
            
            header_fill = PatternFill(start_color="004aad", end_color="004aad", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            border_thick_top = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                      top=Side(style='medium', color="004aad"), bottom=Side(style='thin'))
            
            # Estilizar Header
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border_thin
                
            # Ajustar a largura das colunas
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 60)
            
            # Paleta de cores suaves para cada equipe
            cores_equipe = ["E3F2FD", "E8F5E9", "FFF3E0", "F3E5F5", "FFFDE7", "FCE4EC"]
            equipe_atual = None
            cor_idx = -1
            
            for row in range(2, ws.max_row + 1):
                # Coluna 1 = Equipe Designada, Coluna 2 = Ordem de Execução
                nome_eq = ws.cell(row=row, column=1).value
                
                # Se for uma linha sem equipe (não encontrada)
                if not nome_eq or nome_eq == "":
                    fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    borda = border_thin
                else:
                    if nome_eq != equipe_atual:
                        equipe_atual = nome_eq
                        cor_idx = (cor_idx + 1) % len(cores_equipe)
                        borda = border_thick_top
                    else:
                        borda = border_thin
                        
                    fill = PatternFill(start_color=cores_equipe[cor_idx], end_color=cores_equipe[cor_idx], fill_type="solid")
                
                for col in range(1, ws.max_column + 1):
                    c = ws.cell(row=row, column=col)
                    c.fill = fill
                    c.border = borda
                    if col <= 2: # Centraliza a Equipe e a Ordem
                        c.alignment = Alignment(horizontal="center", vertical="center")

            # Salva o arquivo estilizado
            wb.save(fname)
            
            QMessageBox.information(self, "Sucesso", f"Arquivo salvo e formatado com sucesso!\n{fname}")

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao exportar Excel: {e}")

    def abrir_editar_bairro(self, ponto):
        """Abre o diálogo para editar um ponto existente na rota."""
        # Limpa o nome para edição (remove Zona e Retorno)
        nome_limpo = re.sub(r" \(Zona .*\)", "", ponto['nome'])
        nome_limpo = nome_limpo.replace(" (Retorno)", "")
        
        dialog = AddBairroDialog(nome_limpo, ponto['lat'], ponto['lon'], self)
        if dialog.exec_():
            dados = dialog.get_data()
            
            # 1. Salva no banco de dados (Silent=True para não dar popup toda hora)
            self.salvar_novo_bairro(dados, silent=True)
            
            # 2. Atualiza o objeto em memória (rota atual)
            ponto['lat'] = dados['lat']
            ponto['lon'] = dados['lon']
            ponto['tipo'] = 'manual'
            
            # Recalcula zona e nome
            zona = identificar_zona(dados['lat'], dados['lon'])
            ponto['nome'] = f"{dados['nome']} ({zona})"
            
            # Se for a base (índice 0), atualiza também o ponto de retorno se existir
            if self.rota_atual and self.rota_atual[0] is ponto:
                if len(self.rota_atual) > 1:
                    ultimo = self.rota_atual[-1]
                    if "Retorno" in ultimo['nome']:
                        ultimo['lat'] = dados['lat']
                        ultimo['lon'] = dados['lon']
                        ultimo['nome'] = f"{dados['nome']} ({zona}) (Retorno)"
                        ultimo['tipo'] = 'manual'

            # 3. Atualiza a interface (Mapa e Lista)
            self.atualizar_interface_rota()
            self.txt_log.append(f"Local editado: {dados['nome']}")

    def abrir_adicionar_bairro(self, nome_bairro):
        """Abre o diálogo para adicionar um bairro manualmente."""
        dialog = AddBairroDialog(nome_bairro, parent=self)
        if dialog.exec_():
            dados = dialog.get_data()
            self.salvar_novo_bairro(dados)
            
    def salvar_novo_bairro(self, dados, silent=False):
        """Salva o novo bairro no arquivo JSON oficial."""
        db_data = {}
        # Carrega DB existente
        if os.path.exists(DB_FILE):
            try:
                with open(DB_FILE, 'r', encoding='utf-8') as f:
                    db_data = json.load(f)
            except: pass
            
        # Adiciona novo registro
        db_data[dados['nome']] = {
            "lat": dados['lat'],
            "lon": dados['lon'],
            "tipo": "manual"
        }
        
        # Salva no arquivo
        try:
            with open(DB_FILE, 'w', encoding='utf-8') as f:
                json.dump(db_data, f, ensure_ascii=False, indent=4)
            
            # --- CORREÇÃO CRÍTICA: LIMPAR CACHE ---
            # Se editamos o banco, precisamos remover a entrada antiga do cache
            # para garantir que o roteador use o novo valor do banco.
            if os.path.exists(CACHE_FILE):
                try:
                    with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                        cache = json.load(f)
                    
                    # Remove a entrada se existir (normalizada)
                    chave_norm = normalizar_endereco(dados['nome'])
                    if chave_norm in cache:
                        del cache[chave_norm]
                        with open(CACHE_FILE, 'w', encoding='utf-8') as f:
                            json.dump(cache, f, ensure_ascii=False, indent=4)
                except: pass

            if not silent:
                QMessageBox.information(self, "Sucesso", f"Bairro '{dados['nome']}' salvo com sucesso!")
            self.txt_log.append(f"Banco atualizado: {dados['nome']}")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao salvar no banco de dados: {e}")

    def carregar_excel(self):
        try:
            fname, _ = QFileDialog.getOpenFileName(self, "Abrir Excel", "", "Excel Files (*.xlsx *.xls)")
            if not fname: return
            
            # Lê o arquivo (dtype=str para manter zeros à esquerda em OS/CEP)
            self.df_dados = pd.read_excel(fname, dtype=str, engine='openpyxl')
            # Normaliza colunas (remove espaços e deixa minúsculo para busca)
            self.df_dados.columns = [str(c).strip() for c in self.df_dados.columns]
            
            # Busca coluna de Bairro ou Endereço
            cols = self.df_dados.columns
            col_alvo = next((c for c in cols if 'bairro' in c.lower()), None)
            if not col_alvo:
                col_alvo = next((c for c in cols if 'endereço' in c.lower() or 'endereco' in c.lower()), None)
                
            if col_alvo:
                self.col_endereco_atual = col_alvo
                # Remove duplicatas e vazios
                locais = sorted(list(set([str(x).strip() for x in self.df_dados[col_alvo].dropna().tolist() if str(x).strip()])))
                self.txt_input.setText(", ".join(locais))
                self.txt_log.append(f"Excel carregado! {len(locais)} serviços/bairros encontrados.")
                self.btn_dividir.setEnabled(True)
                self.btn_gerar.setEnabled(True)
            else:
                QMessageBox.warning(self, "Aviso", "Coluna de Bairro/Endereço não encontrada no arquivo.")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao ler Excel: {e}")

    def closeEvent(self, event):
        self.save_settings()
        super().closeEvent(event)

    def save_settings(self):
        try:
            settings = {
                "geometry": self.saveGeometry().toHex().data().decode(),
                "maximized": self.isMaximized()
            }
            with open(SETTINGS_FILE, "w") as f:
                json.dump(settings, f)
        except: pass

    def load_settings(self):
        if os.path.exists(SETTINGS_FILE):
            try:
                with open(SETTINGS_FILE, "r") as f:
                    settings = json.load(f)
                    if "geometry" in settings:
                        self.restoreGeometry(bytes.fromhex(settings["geometry"]))
                    if settings.get("maximized", False):
                        self.setWindowState(self.windowState() | Qt.WindowMaximized)
            except: pass

if __name__ == "__main__":
    app = QApplication(sys.argv)
    janela = LocalizadorApp()
    janela.show()
    sys.exit(app.exec_())
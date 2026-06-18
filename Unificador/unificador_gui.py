# ============================================================================
# UNIFICADOR SAMSYS — Programa Desktop (GUI)
# ============================================================================
# Consolidação e filtragem de OS de Calçadas com interface gráfica moderna.
# Substitui a versão web (Flask) por um programa standalone.
# ============================================================================

import customtkinter as ctk
from tkinter import filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
import threading
import shutil
import json
import io
import os
import sys

# Garante que o console do Windows suporte UTF-8
if sys.stdout:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# ═══════════════════════════════════════════════════════════════════════════
# CONFIGURAÇÃO VISUAL
# ═══════════════════════════════════════════════════════════════════════════
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# Diretório base do programa (onde o .py fica)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(SCRIPT_DIR, "config.json")
BANCO_LOCAL = os.path.join(SCRIPT_DIR, "banco_bairros.xlsx")

# Paleta de cores
COLORS = {
    "bg_dark":      "#0d1117",
    "bg_card":      "#161b22",
    "bg_input":     "#1c2333",
    "border":       "#30363d",
    "accent":       "#3b82f6",
    "accent_hover": "#2563eb",
    "success":      "#10b981",
    "warning":      "#f59e0b",
    "error":        "#ef4444",
    "text":         "#e6edf3",
    "text_sec":     "#8b949e",
    "text_muted":   "#6e7681",
}

# Matriz de códigos de serviço (pavimento e calçada)
CODIGOS_SERVICO = ['613201', '613202', '613203', '613204', '613205', '613206']

# Colunas essenciais que serão mantidas na saída final
COLUNAS_SAIDA = [
    'Número da OS',
    'Data de Solicitação',
    'Bairro',
    'Serviço',
    'Data Limite de Execução',
]


# ═══════════════════════════════════════════════════════════════════════════
# LÓGICA DE PROCESSAMENTO (migrada do app.py)
# ═══════════════════════════════════════════════════════════════════════════

def carregar_arquivo_1(caminho: str) -> pd.DataFrame:
    """Lê Arquivo 1 (Samsys A) — skiprows=4. Coluna: 'Serviço Solicitado'."""
    with open(caminho, 'rb') as f:
        dados = io.BytesIO(f.read())
    df = pd.read_excel(dados, skiprows=4, engine='xlrd')
    df.columns = df.columns.str.strip()
    df['Codigo_Limpo'] = df['Serviço Solicitado'].astype(str).str.strip().str[:6]
    df.rename(columns={'Serviço Solicitado': 'Serviço'}, inplace=True)
    return df


def carregar_arquivo_2(caminho: str) -> pd.DataFrame:
    """Lê Arquivo 2 (Samsys B) — skiprows=4. Código em 'Serviço', descrição em 'Serviço.1'."""
    with open(caminho, 'rb') as f:
        dados = io.BytesIO(f.read())
    df = pd.read_excel(dados, skiprows=4, engine='xlrd')
    df.columns = df.columns.str.strip()
    df['Codigo_Limpo'] = df['Serviço'].astype(str).str.strip().str[:6]

    # Junta código + descrição (ex: "613203 - RECOMPOSIÇÃO DE ASFALTO - LOG")
    if 'Serviço.1' in df.columns:
        codigo = df['Serviço'].astype(str).str.strip()
        descricao = df['Serviço.1'].astype(str).str.strip()
        df['Serviço'] = codigo + ' - ' + descricao
        # Limpa casos onde a descrição era NaN
        df['Serviço'] = df['Serviço'].str.replace(' - nan', '', case=False)

    return df


def extrair_bairros(caminho: str) -> set:
    """Extrai lista de bairros do arquivo norte e oeste.xlsx."""
    with open(caminho, 'rb') as f:
        dados = io.BytesIO(f.read())
    wb = openpyxl.load_workbook(dados, read_only=True, data_only=True)
    ws = wb.active
    bairros = set()

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx == 0:
            continue
        for col_idx in range(0, len(row), 4):
            valor = row[col_idx]
            if valor is not None and isinstance(valor, str):
                nome = valor.strip()
                if nome and nome.upper() != 'BAIRRO':
                    bairros.add(nome.upper())

    wb.close()
    return bairros


def carregar_config() -> dict:
    """Carrega configurações salvas (banco de bairros, etc)."""
    if os.path.isfile(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def salvar_config(cfg: dict):
    """Salva configurações em disco."""
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


# ═══════════════════════════════════════════════════════════════════════════
# COMPONENTE: SLOT DE ARQUIVO
# ═══════════════════════════════════════════════════════════════════════════

class FileSlot(ctk.CTkFrame):
    """Componente reutilizável para seleção de arquivo."""

    def __init__(self, master, label: str, description: str, icon: str,
                 filetypes: list, accent_color: str, **kwargs):
        super().__init__(master, **kwargs)

        self.filepath = None
        self.filetypes = filetypes
        self.on_change_callback = None

        # Config do frame
        self.configure(
            fg_color=COLORS["bg_card"],
            corner_radius=12,
            border_width=1,
            border_color=COLORS["border"],
        )

        # Layout interno
        inner = ctk.CTkFrame(self, fg_color="transparent")
        inner.pack(fill="x", padx=16, pady=14)

        # Linha superior: ícone + label + botão
        top_row = ctk.CTkFrame(inner, fg_color="transparent")
        top_row.pack(fill="x")

        # Ícone
        icon_label = ctk.CTkLabel(
            top_row, text=icon, font=ctk.CTkFont(size=20),
            width=32
        )
        icon_label.pack(side="left", padx=(0, 8))

        # Textos
        text_frame = ctk.CTkFrame(top_row, fg_color="transparent")
        text_frame.pack(side="left", fill="x", expand=True)

        self.title_label = ctk.CTkLabel(
            text_frame, text=label,
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=COLORS["text"],
            anchor="w"
        )
        self.title_label.pack(anchor="w")

        self.desc_label = ctk.CTkLabel(
            text_frame, text=description,
            font=ctk.CTkFont(size=11),
            text_color=COLORS["text_muted"],
            anchor="w"
        )
        self.desc_label.pack(anchor="w")

        # Botão selecionar
        self.btn = ctk.CTkButton(
            top_row, text="Selecionar",
            font=ctk.CTkFont(size=12, weight="bold"),
            width=100, height=32,
            corner_radius=8,
            fg_color=accent_color,
            hover_color=COLORS["accent_hover"],
            command=self._browse
        )
        self.btn.pack(side="right")

        # Status do arquivo (linha inferior)
        self.status_label = ctk.CTkLabel(
            inner, text="Nenhum arquivo selecionado",
            font=ctk.CTkFont(size=11),
            text_color=COLORS["text_muted"],
            anchor="w"
        )
        self.status_label.pack(anchor="w", pady=(8, 0))

    def _browse(self):
        path = filedialog.askopenfilename(filetypes=self.filetypes)
        if path:
            self.filepath = path
            filename = os.path.basename(path)
            self.status_label.configure(
                text=f"✅  {filename}",
                text_color=COLORS["success"]
            )
            self.configure(border_color=COLORS["success"])
            if self.on_change_callback:
                self.on_change_callback()

    def set_on_change(self, callback):
        self.on_change_callback = callback

    def get_path(self):
        return self.filepath

    def set_path(self, path: str):
        """Define o caminho programaticamente (ex: auto-load do banco salvo)."""
        if path and os.path.isfile(path):
            self.filepath = path
            filename = os.path.basename(path)
            self.status_label.configure(
                text=f"✅  {filename}",
                text_color=COLORS["success"]
            )
            self.configure(border_color=COLORS["success"])
            if self.on_change_callback:
                self.on_change_callback()


# ═══════════════════════════════════════════════════════════════════════════
# APLICAÇÃO PRINCIPAL
# ═══════════════════════════════════════════════════════════════════════════

class UnificadorApp(ctk.CTk):
    """Janela principal do Unificador Samsys."""

    def __init__(self):
        super().__init__()

        # Janela
        self.title("Unificador Samsys")
        self.geometry("720x680")
        self.minsize(620, 580)
        self.configure(fg_color=COLORS["bg_dark"])

        # Centralizar na tela
        self.update_idletasks()
        w = 720
        h = 680
        x = (self.winfo_screenwidth() // 2) - (w // 2)
        y = (self.winfo_screenheight() // 2) - (h // 2)
        self.geometry(f"{w}x{h}+{x}+{y}")

        self._processing = False
        self._build_ui()
        self._load_banco_bairros()

    def _build_ui(self):
        # ── Container principal com scroll ──
        container = ctk.CTkFrame(self, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=24, pady=20)

        # ── Header ──
        header = ctk.CTkFrame(container, fg_color="transparent")
        header.pack(fill="x", pady=(0, 8))

        # Badge
        badge = ctk.CTkLabel(
            header, text="  ● Automação Samsys",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color=COLORS["success"],
            anchor="center"
        )
        badge.pack()

        # Título
        title = ctk.CTkLabel(
            header, text="Unificador de Planilhas",
            font=ctk.CTkFont(size=26, weight="bold"),
            text_color=COLORS["text"]
        )
        title.pack(pady=(6, 2))

        # Subtítulo
        subtitle = ctk.CTkLabel(
            header,
            text="Consolide relatórios, filtre por pavimento e calçada, exporte a base limpa.",
            font=ctk.CTkFont(size=12),
            text_color=COLORS["text_sec"]
        )
        subtitle.pack(pady=(0, 16))

        # Separador fino
        sep = ctk.CTkFrame(container, fg_color=COLORS["border"], height=1)
        sep.pack(fill="x", pady=(0, 16))

        # ── Slots de Arquivo ──
        xls_types = [("Planilha Excel", "*.xls *.xlsx"), ("Todos", "*.*")]
        xlsx_types = [("Planilha Excel", "*.xlsx"), ("Todos", "*.*")]

        self.slot1 = FileSlot(
            container,
            label="Samsys A",
            description="Relatório com 4 linhas de cabeçalho (.xls)",
            icon="📄",
            filetypes=xls_types,
            accent_color=COLORS["accent"],
        )
        self.slot1.pack(fill="x", pady=(0, 8))

        self.slot2 = FileSlot(
            container,
            label="Samsys B",
            description="Relatório com 4 linhas de cabeçalho (.xls)",
            icon="📄",
            filetypes=xls_types,
            accent_color="#06b6d4",
        )
        self.slot2.pack(fill="x", pady=(0, 8))

        self.slot3 = FileSlot(
            container,
            label="Banco de Bairros",
            description="Planilha norte e oeste (.xlsx)",
            icon="📍",
            filetypes=xlsx_types,
            accent_color=COLORS["warning"],
        )
        self.slot3.pack(fill="x", pady=(0, 16))

        # Conectar callbacks para verificar prontidão
        for slot in [self.slot1, self.slot2, self.slot3]:
            slot.set_on_change(self._check_ready)

        # Callback extra no slot3: salvar + copiar banco de bairros
        original_browse = self.slot3._browse
        def _browse_and_save():
            original_browse()
            if self.slot3.get_path():
                self._save_banco_bairros(self.slot3.get_path())
        self.slot3._browse = _browse_and_save

        # ── Botão Processar ──
        self.btn_process = ctk.CTkButton(
            container,
            text="⚡  Processar Planilhas",
            font=ctk.CTkFont(size=15, weight="bold"),
            height=46,
            corner_radius=10,
            fg_color=COLORS["accent"],
            hover_color=COLORS["accent_hover"],
            state="disabled",
            command=self._on_process
        )
        self.btn_process.pack(fill="x", pady=(0, 16))

        # ── Barra de Progresso ──
        self.progress = ctk.CTkProgressBar(
            container,
            height=6,
            corner_radius=3,
            fg_color=COLORS["bg_card"],
            progress_color=COLORS["accent"],
        )
        self.progress.pack(fill="x", pady=(0, 12))
        self.progress.set(0)

        # ── Console / Log ──
        console_frame = ctk.CTkFrame(
            container,
            fg_color=COLORS["bg_card"],
            corner_radius=12,
            border_width=1,
            border_color=COLORS["border"],
        )
        console_frame.pack(fill="both", expand=True)

        # Header do console
        console_header = ctk.CTkFrame(console_frame, fg_color="#0d1117", corner_radius=0, height=32)
        console_header.pack(fill="x")
        console_header.pack_propagate(False)

        # Bolinhas estilo terminal
        dots_frame = ctk.CTkFrame(console_header, fg_color="transparent")
        dots_frame.pack(side="left", padx=12, pady=8)

        for color in ["#ef4444", "#f59e0b", "#10b981"]:
            dot = ctk.CTkFrame(dots_frame, width=10, height=10, corner_radius=5, fg_color=color)
            dot.pack(side="left", padx=2)

        console_title = ctk.CTkLabel(
            console_header, text="Console",
            font=ctk.CTkFont(size=11),
            text_color=COLORS["text_muted"]
        )
        console_title.pack(side="left", padx=8)

        # Área de texto do console
        self.console = ctk.CTkTextbox(
            console_frame,
            font=ctk.CTkFont(family="Consolas", size=12),
            fg_color=COLORS["bg_input"],
            text_color=COLORS["text_sec"],
            corner_radius=0,
            border_width=0,
            state="disabled",
            wrap="word",
        )
        self.console.pack(fill="both", expand=True, padx=1, pady=(0, 1))

        # Tag para linhas de destaque
        self.console._textbox.tag_config("success", foreground=COLORS["success"])
        self.console._textbox.tag_config("error", foreground=COLORS["error"])
        self.console._textbox.tag_config("info", foreground=COLORS["accent"])
        self.console._textbox.tag_config("warning", foreground=COLORS["warning"])

    # ───────────────────────────────────────────────────────────
    # BANCO DE BAIRROS (persistência)
    # ───────────────────────────────────────────────────────────

    def _save_banco_bairros(self, caminho: str):
        """Salva o caminho do banco de bairros e copia para a pasta base."""
        try:
            # Copia o arquivo para a pasta base como backup
            if os.path.abspath(caminho) != os.path.abspath(BANCO_LOCAL):
                shutil.copy2(caminho, BANCO_LOCAL)

            # Salva o caminho no config
            cfg = carregar_config()
            cfg['banco_bairros'] = BANCO_LOCAL
            salvar_config(cfg)
        except Exception:
            pass

    def _load_banco_bairros(self):
        """Carrega automaticamente o banco de bairros salvo."""
        cfg = carregar_config()
        caminho = cfg.get('banco_bairros', '')

        # Verifica se o banco local existe
        if os.path.isfile(BANCO_LOCAL):
            self.slot3.set_path(BANCO_LOCAL)
        elif caminho and os.path.isfile(caminho):
            self.slot3.set_path(caminho)

    # ───────────────────────────────────────────────────────────
    # HELPERS
    # ───────────────────────────────────────────────────────────

    def _check_ready(self):
        """Habilita o botão quando os 3 arquivos estão selecionados."""
        all_ready = all([
            self.slot1.get_path(),
            self.slot2.get_path(),
            self.slot3.get_path(),
        ])
        if all_ready and not self._processing:
            self.btn_process.configure(state="normal")

    def _log(self, msg: str, tag: str = ""):
        """Adiciona uma linha ao console."""
        self.console.configure(state="normal")
        if tag:
            self.console._textbox.insert("end", f"> {msg}\n", tag)
        else:
            self.console.insert("end", f"> {msg}\n")
        self.console.configure(state="disabled")
        self.console.see("end")

    def _set_progress(self, value: float):
        """Atualiza a barra de progresso (0.0 a 1.0)."""
        self.progress.set(value)

    # ───────────────────────────────────────────────────────────
    # PROCESSAMENTO
    # ───────────────────────────────────────────────────────────

    def _on_process(self):
        """Inicia o processamento em uma thread separada."""
        if self._processing:
            return

        self._processing = True
        self.btn_process.configure(state="disabled", text="⏳  Processando...")

        # Limpa console
        self.console.configure(state="normal")
        self.console.delete("1.0", "end")
        self.console.configure(state="disabled")
        self._set_progress(0)

        # Thread de processamento
        thread = threading.Thread(target=self._process_worker, daemon=True)
        thread.start()

    def _process_worker(self):
        """Worker que roda em background para não travar a UI."""
        try:
            path1 = self.slot1.get_path()
            path2 = self.slot2.get_path()
            path3 = self.slot3.get_path()

            # ── Validação de arquivos ──
            nomes = {"Samsys A": path1, "Samsys B": path2, "Banco de Bairros": path3}
            for nome, caminho in nomes.items():
                if not caminho or not os.path.isfile(caminho):
                    raise ValueError(f"Arquivo não encontrado: {nome}")
                tamanho = os.path.getsize(caminho)
                if tamanho == 0:
                    raise ValueError(
                        f"Arquivo '{nome}' está vazio (0 bytes) — "
                        f"provavelmente corrompido. Baixe novamente: {os.path.basename(caminho)}"
                    )

            # ── Etapa 1: Ler Arquivo 1 ──
            self.after(0, self._log, f"Lendo Samsys A: {os.path.basename(path1)}...", "info")
            df1 = carregar_arquivo_1(path1)
            self.after(0, self._log, f"  → {len(df1)} registros, {len(df1.columns)} colunas.")
            self.after(0, self._set_progress, 0.15)

            # ── Etapa 2: Ler Arquivo 2 ──
            self.after(0, self._log, f"Lendo Samsys B: {os.path.basename(path2)}...", "info")
            df2 = carregar_arquivo_2(path2)
            self.after(0, self._log, f"  → {len(df2)} registros, {len(df2.columns)} colunas.")
            self.after(0, self._set_progress, 0.30)

            # ── Etapa 3: Concatenação ──
            self.after(0, self._log, "Concatenando dataframes...", "info")
            df = pd.concat([df1, df2], ignore_index=True)
            df.columns = df.columns.str.strip()
            self.after(0, self._log, f"  → Base consolidada: {len(df)} registros, {len(df.columns)} colunas.")
            self.after(0, self._set_progress, 0.45)

            # ── Etapa 4: Filtro de serviços ──
            self.after(0, self._log, "Filtrando por códigos de pavimento e calçada...", "info")
            df_filtrado = df[df['Codigo_Limpo'].isin(CODIGOS_SERVICO)].copy()
            descartados = len(df) - len(df_filtrado)
            self.after(0, self._log, f"  → {len(df_filtrado)} retidos ({descartados} descartados por serviço).")
            self.after(0, self._set_progress, 0.60)

            # ── Etapa 5: Extrair e filtrar bairros ──
            self.after(0, self._log, f"Extraindo banco de bairros: {os.path.basename(path3)}...", "info")
            bairros_validos = extrair_bairros(path3)
            self.after(0, self._log, f"  → {len(bairros_validos)} bairros únicos extraídos.")

            self.after(0, self._log, "Filtrando por bairros (norte e oeste)...", "info")
            df_filtrado['_bairro_norm'] = df_filtrado['Bairro'].astype(str).str.strip().str.upper()
            df_final = df_filtrado[df_filtrado['_bairro_norm'].isin(bairros_validos)].copy()
            descartados_bairro = len(df_filtrado) - len(df_final)
            df_final.drop(columns=['_bairro_norm', 'Codigo_Limpo'], inplace=True)
            self.after(0, self._log, f"  → {len(df_final)} retidos ({descartados_bairro} descartados por bairro).")
            self.after(0, self._set_progress, 0.80)

            # ── Etapa 6: Selecionar colunas ──
            self.after(0, self._log, "Selecionando colunas essenciais...", "info")
            colunas_presentes = [c for c in COLUNAS_SAIDA if c in df_final.columns]
            colunas_faltando = [c for c in COLUNAS_SAIDA if c not in df_final.columns]

            if colunas_faltando:
                self.after(0, self._log, f"  ⚠ Colunas não encontradas: {', '.join(colunas_faltando)}", "warning")

            df_final = df_final[colunas_presentes]
            self.after(0, self._log, f"  → Colunas mantidas: {', '.join(colunas_presentes)}")
            self.after(0, self._set_progress, 0.90)

            # ── Salvar ──
            # Pedir local para salvar (precisa rodar na thread principal)
            self.after(0, self._ask_save, df_final)

        except Exception as e:
            self.after(0, self._log, f"ERRO: {str(e)}", "error")
            self.after(0, self._reset_button)

    def _ask_save(self, df_final: pd.DataFrame):
        """Abre o diálogo Salvar Como na thread principal."""
        hoje = date.today().strftime("%Y-%m-%d")
        nome_padrao = f"Unificado_Samsys_{hoje}.xlsx"

        save_path = filedialog.asksaveasfilename(
            title="Salvar planilha processada",
            defaultextension=".xlsx",
            initialfile=nome_padrao,
            filetypes=[("Planilha Excel", "*.xlsx"), ("Todos", "*.*")]
        )

        if save_path:
            try:
                df_final.to_excel(save_path, index=False, engine='openpyxl')

                # ── Estilizar o Excel ──
                wb = openpyxl.load_workbook(save_path)
                ws = wb.active

                # Estilos
                header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
                header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell_font = Font(name='Calibri', size=10)
                cell_align = Alignment(horizontal='left', vertical='center', wrap_text=False)
                thin_border = Border(
                    left=Side(style='thin', color='D1D5DB'),
                    right=Side(style='thin', color='D1D5DB'),
                    top=Side(style='thin', color='D1D5DB'),
                    bottom=Side(style='thin', color='D1D5DB'),
                )
                alt_fill = PatternFill(start_color='F0F4FF', end_color='F0F4FF', fill_type='solid')

                # Aplicar estilo ao cabeçalho
                for col_idx, cell in enumerate(ws[1], 1):
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                    cell.border = thin_border

                # Aplicar estilo às linhas de dados + zebra striping
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column), 2):
                    for cell in row:
                        cell.font = cell_font
                        cell.alignment = cell_align
                        cell.border = thin_border
                        if row_idx % 2 == 0:
                            cell.fill = alt_fill

                # Auto-fit: ajustar largura das colunas ao conteúdo
                for col_idx in range(1, ws.max_column + 1):
                    max_len = 0
                    col_letter = get_column_letter(col_idx)
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            if cell.value is not None:
                                cell_len = len(str(cell.value))
                                if cell_len > max_len:
                                    max_len = cell_len
                    # Adiciona margem + limita largura máxima
                    ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

                # Congelar painel no cabeçalho
                ws.freeze_panes = 'A2'

                wb.save(save_path)
                wb.close()

                self._set_progress(1.0)
                self._log(f"", "")
                self._log(f"✅ Exportado com sucesso!", "success")
                self._log(f"   {len(df_final)} registros | {len(df_final.columns)} colunas", "success")
                self._log(f"   Salvo em: {save_path}", "success")
                self.progress.configure(progress_color=COLORS["success"])
            except Exception as e:
                self._log(f"ERRO ao salvar: {str(e)}", "error")
        else:
            self._log("Exportação cancelada pelo usuário.", "warning")

        self._reset_button()

    def _reset_button(self):
        """Restaura o botão ao estado original."""
        self._processing = False
        self.btn_process.configure(text="⚡  Processar Planilhas")
        self.progress.configure(progress_color=COLORS["accent"])
        self._check_ready()


# ═══════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    app = UnificadorApp()
    app.mainloop()

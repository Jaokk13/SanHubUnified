import os
import io
import pandas as pd
from datetime import date
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Body
from fastapi.responses import JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import database as db
import router_engine as router

app = FastAPI(title="SanHub Unified API")

# Initialize database
db.init_db()

# Mount static files
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(SCRIPT_DIR, "static")
TEMPLATES_DIR = os.path.join(SCRIPT_DIR, "templates")

if not os.path.exists(STATIC_DIR):
    os.makedirs(os.path.join(STATIC_DIR, "css"))
    os.makedirs(os.path.join(STATIC_DIR, "js"))
if not os.path.exists(TEMPLATES_DIR):
    os.makedirs(TEMPLATES_DIR)

app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

CODIGOS_SERVICO = ['613201', '613202', '613203', '613204', '613205', '613206']

@app.get("/")
def read_root():
    return FileResponse(os.path.join(TEMPLATES_DIR, "index.html"))

# ─────────────────────────────────────────────────────────────────────────────
# IMPORTAÇÃO SAMSYS
# ─────────────────────────────────────────────────────────────────────────────

def _extrair_bairros_validos(file_content: bytes) -> set:
    wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    ws = wb.active
    bairros = set()
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx == 0:
            continue
        for col_idx in range(0, len(row), 4):
            valor = row[col_idx]
            if valor is not None and isinstance(valor, str):
                nome = valor.strip()
                if nome and nome.upper() != 'BAIRRO':
                    bairros.add(nome.upper())
    wb.close()
    return bairros

def _determinar_categoria(descricao: str) -> str:
    desc = str(descricao).upper()
    
    # Check for Calçada keywords first
    if any(p in desc for p in ["CALÇADA", "CALCADA", "PASSEIO", "GUIA", "MEIO-FIO", "CIMENTADO", "LAJOTA"]):
        return "Calçada"
        
    # Then check for Asfalto keywords
    if any(p in desc for p in ["ASFALTO", "PAVIMENTO", "CBUQ", "TAPA", "BURACO", "RECOMPOSIÇÃO DE ASFALTO"]):
        return "Asfalto"
        
    return "Indefinido"

@app.post("/api/import")
async def import_samsys(
    file_a: UploadFile = File(...),
    file_b: UploadFile = File(...),
    file_bairros: Optional[UploadFile] = File(None)
):
    try:
        # Lendo Arquivo A
        content_a = await file_a.read()
        df1 = pd.read_excel(io.BytesIO(content_a), skiprows=4, engine='xlrd' if file_a.filename.endswith('.xls') else 'openpyxl')
        df1.columns = df1.columns.str.strip()
        df1['Codigo_Limpo'] = df1['Serviço Solicitado'].astype(str).str.strip().str[:6]
        df1.rename(columns={'Serviço Solicitado': 'Serviço'}, inplace=True)

        # Lendo Arquivo B
        content_b = await file_b.read()
        df2 = pd.read_excel(io.BytesIO(content_b), skiprows=4, engine='xlrd' if file_b.filename.endswith('.xls') else 'openpyxl')
        df2.columns = df2.columns.str.strip()
        df2['Codigo_Limpo'] = df2['Serviço'].astype(str).str.strip().str[:6]
        if 'Serviço.1' in df2.columns:
            codigo = df2['Serviço'].astype(str).str.strip()
            descricao = df2['Serviço.1'].astype(str).str.strip()
            df2['Serviço'] = codigo + ' - ' + descricao
            df2['Serviço'] = df2['Serviço'].str.replace(' - nan', '', case=False)

        # Concatenação e filtro por serviços
        df = pd.concat([df1, df2], ignore_index=True)
        df.columns = df.columns.str.strip()
        df_filtrado = df[df['Codigo_Limpo'].isin(CODIGOS_SERVICO)].copy()

        # Filtro de Bairros (Opcional, se o arquivo foi enviado)
        if file_bairros:
            content_bairros = await file_bairros.read()
            bairros_validos = _extrair_bairros_validos(content_bairros)
            df_filtrado['_bairro_norm'] = df_filtrado['Bairro'].astype(str).str.strip().str.upper()
            df_filtrado = df_filtrado[df_filtrado['_bairro_norm'].isin(bairros_validos)].copy()

        # Encontrando colunas importantes
        col_os = next((c for c in df_filtrado.columns if c.lower().strip() in ['número da os', 'numero da os', 'os', 'o.s']), None)
        col_data_sol = next((c for c in df_filtrado.columns if 'solicitação' in c.lower() or 'solicitacao' in c.lower()), None)
        col_bairro = next((c for c in df_filtrado.columns if 'bairro' in c.lower()), None)
        col_endereco = next((c for c in df_filtrado.columns if c.lower().strip() in ['endereço', 'endereco', 'logradouro', 'rua']), None)
        col_servico = 'Serviço'
        col_data_limite = next((c for c in df_filtrado.columns if 'limite' in c.lower()), None)
        
        # Match exato para não pegar "Situação Ligação Água" nem "Parecer Solicitação"
        col_situacao = next((c for c in df_filtrado.columns if c.lower().strip() in ['situação', 'situacao', 'status']), None)
        col_parecer = next((c for c in df_filtrado.columns if c.lower().strip() in ['parecer não execução', 'parecer nao execucao', 'motivo não execução']), None)

        if not col_os:
            raise HTTPException(status_code=400, detail="Coluna de Número da OS não encontrada.")

        orders_to_insert = []
        current_os_numbers = set()

        for _, row in df_filtrado.iterrows():
            os_number = str(row[col_os]).strip()
            current_os_numbers.add(os_number)
            servico_desc = str(row[col_servico]) if col_servico in row else ""
            situacao_val = str(row[col_situacao]).upper() if col_situacao in row else ""
            parecer_val = str(row[col_parecer]) if col_parecer in row else ""
            
            is_post = "POSTERGADO" in situacao_val or "POSTERGADA" in situacao_val
            postergo_reason = parecer_val if is_post else ""
            
            # Se a OS for Cortada, ela eh considerada ativa (is_postergada = False)
            if is_post and postergo_reason:
                is_cortada = bool(re.search(r'\d+[,.]?\d*\s*(?:m|mts|cm)?\s*[xX]\s*\d+[,.]?\d*\s*(?:m|mts|cm)?', postergo_reason, re.IGNORECASE))
                if is_cortada:
                    is_post = False

            orders_to_insert.append({
                "os_number": os_number,
                "solicitation_date": str(row[col_data_sol]) if col_data_sol in row else "",
                "neighborhood": str(row[col_bairro]) if col_bairro in row else "",
                "address": str(row[col_endereco]) if col_endereco in row else "",
                "service_description": servico_desc,
                "limit_date": str(row[col_data_limite]) if col_data_limite in row else "",
                "category": _determinar_categoria(servico_desc),
                "is_postergada": is_post,
                "postergo_reason": postergo_reason
            })

        # Insere novas OS
        insert_res = db.upsert_orders(orders_to_insert)
        
        # Reconciliação: Marca como executado o que sumiu
        exec_count = db.reconcile_executed(current_os_numbers)

        return {
            "success": True,
            "message": f"Importação concluída. {insert_res['inserted']} novas OS cadastradas. {exec_count} OS marcadas como executadas."
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ─────────────────────────────────────────────────────────────────────────────
# APIS GERAIS
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/stats")
def get_stats():
    return db.get_stats()

@app.get("/api/stats/chart")
def get_chart_stats():
    return db.get_chart_stats()

@app.get("/api/orders")
def get_orders(status: Optional[str] = None, category: Optional[str] = None, team_id: Optional[int] = None, search: Optional[str] = None, date: Optional[str] = None):
    return db.get_orders(status, category, team_id, search, date)

class AssignRequest(BaseModel):
    os_numbers: List[str]
    team_id: Optional[int] = None
    date: Optional[str] = None

@app.post("/api/orders/assign")
def assign_orders(req: AssignRequest):
    for os_num in req.os_numbers:
        db.assign_order_to_team(os_num, req.team_id, req.date)
    return {"success": True}

class ReorderRequest(BaseModel):
    team_id: int
    os_numbers: List[str]

@app.post("/api/orders/reorder")
def reorder_orders(req: ReorderRequest):
    db.set_execution_orders(req.team_id, req.os_numbers)
    return {"success": True}

class UpdateStateRequest(BaseModel):
    state: str

@app.post("/api/orders/{os_number}/state")
def update_os_state(os_number: str, req: UpdateStateRequest):
    db.update_os_state(os_number, req.state)
    return {"success": True}

class AutoAssignRequest(BaseModel):
    category: str
    date: Optional[str] = None
    task_type: str
    max_orders: Optional[int] = None
    team_ids: Optional[List[int]] = None
    specific_os_list: Optional[str] = None

@app.post("/api/orders/auto-assign")
def auto_assign_orders(req: AutoAssignRequest):
    # 1. Obter todas as equipes dessa categoria e dessa função (Prévia ou Execução)
    teams = [t for t in db.get_teams() if t["type"] == req.category and t["task_type"] == req.task_type]
    
    # 1.5 Filtrar apenas as equipes selecionadas pelo usuário (se enviado)
    if req.team_ids is not None:
        teams = [t for t in teams if t["id"] in req.team_ids]
        
    if not teams:
        raise HTTPException(status_code=400, detail="Nenhuma equipe cadastrada ou selecionada para esta Categoria e Função.")
    
    # 2. Obter todas as OSs pendentes dessa categoria sem equipe
    all_orders = db.get_orders(status="Pendente", category=req.category, team_id=None)
    
    # 3. Filtrar as OSs baseadas na Função
    # Se Função for Execução, precisamos de OSs Cortadas (com medidas).
    # Se Função for Prévia, precisamos de OSs Não-Cortadas (sem medidas).
    orders = []
    import re
    for o in all_orders:
        if o.get('is_postergada'):
            continue  # Ignorar OS postergadas no Programador Automático
            
        if req.category == 'Calçada':
            # Para calçada, todas as OS são consideradas como Execução
            if req.task_type == 'Execução':
                orders.append(o)
        else:
            is_cortada = False
            if o.get('is_postergada') and o.get('postergo_reason'):
                is_cortada = bool(re.search(r'\d+[,.]?\d*\s*(?:m|mts|cm)?\s*[xX]\s*\d+[,.]?\d*\s*(?:m|mts|cm)?', o['postergo_reason'], re.IGNORECASE))
            
            if req.task_type == 'Execução' and is_cortada:
                orders.append(o)
            elif req.task_type == 'Prévia' and not is_cortada:
                orders.append(o)

    # Filtro adicional por lista específica de OS
    if req.specific_os_list:
        specific_os_set = {os.strip() for os in req.specific_os_list.split(',')}
        orders = [o for o in orders if str(o['os_number']).strip() in specific_os_set]

    if not orders:
        raise HTTPException(status_code=400, detail="Nenhuma OS pendente sem equipe se encaixa nesta Função (Prévia/Execução) ou na lista fornecida.")
    # 4. Aplicar algoritmo Sweep
    check_mass = (req.category == 'Asfalto' and req.task_type == 'Execução')
    divisao = router.dividir_em_equipes_sweep(orders, len(teams), check_mass=check_mass, max_orders=req.max_orders)
    
    # 5. Atribuir OSs às equipes
    for i, team in enumerate(teams):
        if i in divisao:
            for os_num in divisao[i]:
                db.assign_order_to_team(os_num, team["id"], req.date)
                
    return {"success": True, "message": f"Divisão automática via Sweep concluída para {len(orders)} OS(s)!"}

@app.get("/api/teams")
def get_teams():
    return db.get_teams()

class TeamCreate(BaseModel):
    name: str
    type: str
    task_type: str = 'Execução'

@app.post("/api/teams")
def create_team(team: TeamCreate):
    try:
        return db.create_team(team.name, team.type, team.task_type)
    except Exception as e:
        if "UNIQUE constraint failed" in str(e):
            raise HTTPException(status_code=400, detail="Já existe uma equipe com este nome.")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/teams/{team_id}")
def update_team(team_id: int, team: TeamCreate):
    try:
        db.update_team(team_id, team.name, team.type, team.task_type)
        return {"success": True}
    except Exception as e:
        if "UNIQUE constraint failed" in str(e):
            raise HTTPException(status_code=400, detail="Já existe uma equipe com este nome.")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/teams/{team_id}")
def delete_team(team_id: int):
    db.delete_team(team_id)
    return {"success": True}

# ─────────────────────────────────────────────────────────────────────────────
# ROTAS E EXPORTAÇÃO
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/route/{team_id}")
def calculate_route(team_id: int, date: Optional[str] = None):
    import traceback
    try:
        orders = db.get_orders(status="Pendente", team_id=team_id, scheduled_date=date)
        if not orders:
            raise HTTPException(status_code=400, detail="Nenhuma OS encontrada para esta equipe.")
        res = router.roteirizar_equipe(orders)
        
        # Test JSON serialization right here to see if it fails
        import json
        json.dumps(res)
        
        # Não salva mais automaticamente no banco
        # Apenas retorna a rota calculada para o frontend
        return res
    except Exception as e:
        err_str = traceback.format_exc()
        raise HTTPException(status_code=500, detail=f"Erro interno: {err_str}")

@app.post("/api/route/{team_id}/confirm")
def confirm_route(team_id: int, os_list: list[str] = Body(...)):
    try:
        db.set_execution_orders(team_id, os_list)
        return {"success": True, "message": "Rota confirmada e salva com sucesso!"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/route/{team_id}")
def get_saved_route(team_id: int, date: Optional[str] = None):
    try:
        orders = db.get_orders(status="Pendente", team_id=team_id, scheduled_date=date)
        if not orders:
            raise HTTPException(status_code=404, detail="Nenhuma OS para esta equipe.")
            
        # Verifica se pelo menos uma tem execution_order
        has_route = any(o.get("execution_order") is not None for o in orders)
        if not has_route:
            raise HTTPException(status_code=404, detail="Rota ainda não confirmada para esta equipe.")
            
        res = router.montar_rota_salva(orders)
        return res
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/cache")
def clear_geocoding_cache():
    db.clear_cache()
    return {"success": True, "message": "Cache temporário apagado com sucesso."}

class ResetRoutesRequest(BaseModel):
    date: str

@app.post("/api/reset-routes")
def reset_routes(req: ResetRoutesRequest):
    try:
        db.reset_past_routes(req.date)
        return {"success": True, "message": f"Rotas apagadas para a data {req.date}."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/settings")
def get_settings():
    return db.get_settings()

@app.post("/api/settings")
def save_settings(data: dict = Body(...)):
    base_lat = data.get("base_lat")
    base_lon = data.get("base_lon")
    if base_lat:
        db.save_setting("base_lat", str(base_lat))
    if base_lon:
        db.save_setting("base_lon", str(base_lon))
    return {"status": "ok", "message": "Configurações salvas com sucesso!"}

@app.get("/api/export")
def export_excel(type: str = "todas", date: str = None):
    # Pega todas as OS
    orders = db.get_orders()

    if type == "programadas":
        orders = [o for o in orders if o.get('team_id') is not None and o.get('status') == 'Pendente']
        if date:
            orders = [o for o in orders if o.get('scheduled_date') == date]
    elif type == "executadas":
        orders = [o for o in orders if o.get('status') == 'Executado']

    if not orders:
        raise HTTPException(status_code=400, detail="Nenhum dado para exportar com os filtros informados.")
        
    df = pd.DataFrame(orders)
    
    # Formatação e nomes de colunas
    df['Equipe Designada'] = df['team_name'].fillna("Não Atribuída")
    df['Função (Equipe)'] = df.apply(lambda row: row.get('team_task_type') if 'team_task_type' in row else '', axis=1)
    df['Ordem de Execução'] = df['execution_order']
    df['Status'] = df['status']
    df['Número da OS'] = df['os_number']
    df['Bairro'] = df['neighborhood']
    df['Endereço'] = df['address'] if 'address' in df.columns else ''
    df['Serviço'] = df['service_description']
    df['Categoria'] = df['category']
    df['Data Limite'] = df['limit_date']
    
    cols_to_keep = [
        'Equipe Designada', 'Função (Equipe)', 'Ordem de Execução', 'Status', 'Número da OS',
        'Bairro', 'Endereço', 'Serviço', 'Categoria', 'Data Limite'
    ]
    df = df[cols_to_keep]
    df = df.sort_values(by=['Status', 'Equipe Designada', 'Ordem de Execução'], ascending=[False, True, True])
    
    filepath = os.path.join(SCRIPT_DIR, "Rota_Equipes.xlsx")
    df.to_excel(filepath, index=False)
    
    # Estilização
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    header_fill = PatternFill(start_color="004aad", end_color="004aad", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
                         
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin
        
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)
        
    # Zebra striping
    alt_fill = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        for cell in row:
            cell.border = border_thin
            if row_idx % 2 == 0:
                cell.fill = alt_fill
                
    wb.save(filepath)
    return FileResponse(filepath, filename="Planejamento_OS.xlsx")

@app.get("/api/cache/export")
def export_cache():
    import json
    local_db_path = os.path.join(SCRIPT_DIR, "banco_bairros_compartilhar.json")
    if not os.path.exists(local_db_path):
        # Create empty if not exists
        with open(local_db_path, "w", encoding="utf-8") as f:
            json.dump({}, f)
    return FileResponse(local_db_path, filename="banco_bairros_compartilhar.json", media_type="application/json")

@app.post("/api/cache/import")
async def import_cache(file: UploadFile = File(...)):
    import json
    try:
        content = await file.read()
        data = json.loads(content)
        
        # Validação básica
        if not isinstance(data, dict):
            raise ValueError("O arquivo JSON não é um dicionário válido de bairros.")
            
        local_db_path = os.path.join(SCRIPT_DIR, "banco_bairros_compartilhar.json")
        
        # Merge com o existente (opcional: ou sobrescrever)
        existing = {}
        if os.path.exists(local_db_path):
            with open(local_db_path, "r", encoding="utf-8") as f:
                try: existing = json.load(f)
                except: pass
        
        existing.update(data)
        
        with open(local_db_path, "w", encoding="utf-8") as f:
            json.dump(existing, f, ensure_ascii=False, indent=4)
            
        # Invalida o cache em memória do router_engine
        router._local_db_cache = None
        
        return {"success": True, "message": "Banco de bairros atualizado com sucesso!"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao importar JSON: {e}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
